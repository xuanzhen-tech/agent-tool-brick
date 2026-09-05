"""AgentTool 表格工具的确定性 Python worker。

本文件只接受 JSON 声明并执行受控的数据读取、聚合和校验。它不执行模型提供的
Python、SQL、Excel 公式或任意表达式，所有金额和比率运算统一使用 Decimal。
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import sys
import uuid
import zipfile
from collections import Counter, defaultdict, deque
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_EVEN, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


MAX_SOURCE_BYTES = 100 * 1024 * 1024
MAX_SOURCES = 100
MAX_TOTAL_SOURCE_BYTES = 1024 * 1024 * 1024
MAX_ARCHIVE_BYTES = 512 * 1024 * 1024
MAX_ARCHIVE_RATIO = 200
MAX_NON_EMPTY_CELLS = 2_000_000
MAX_SCANNED_CELLS = 5_000_000
MAX_TABLE_COLUMNS = 512
MAX_JOIN_OUTPUT_ROWS = 1_000_000
MAX_QUERY_ROWS = 100_000
MAX_QUERIES = 20
MAX_CHECKS = 100
MAX_NOT_COMPUTABLE_SAMPLES = 100
MAX_PUBLIC_NOT_COMPUTABLE_SAMPLES = 5
MAX_PUBLIC_PREVIEW_ROWS = 10
MAX_PUBLIC_PREVIEW_CELLS = 60
MAX_PUBLIC_PREVIEW_COLUMNS = 20
MAX_PUBLIC_VALIDATION_CHECKS = 20
MAX_PUBLIC_EVIDENCE_ITEMS = 20
MAX_PUBLIC_EVIDENCE_FIELDS = 40
MAX_PUBLIC_EVIDENCE_STRING_CHARS = 500
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".csv", ".tsv"}
ID_RE = re.compile(r"^(?:analysis|result|validation)-[a-z0-9-]{8,80}$")


class SpreadsheetError(Exception):
    """携带稳定错误码的预期业务异常。"""

    def __init__(self, code: str, message: str, details: dict[str, Any] | None = None, blocked: bool = False):
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = details or {}
        self.blocked = blocked


def main() -> None:
    try:
        request = json.load(sys.stdin)
        workspace = resolve_workspace(request.get("workspace"))
        arguments = request.get("arguments") or {}
        if not isinstance(arguments, dict):
            raise SpreadsheetError("spreadsheet_invalid_input", "工具参数必须是对象。", blocked=True)
        action = request.get("action")
        if action == "inspect":
            value = inspect_spreadsheet(workspace, arguments)
        elif action == "compute":
            value = compute_spreadsheet(workspace, arguments)
        elif action == "validate":
            value = validate_spreadsheet(workspace, arguments)
        else:
            raise SpreadsheetError("spreadsheet_action_invalid", f"未知表格操作: {action}", blocked=True)
        emit({"ok": True, "value": value})
    except SpreadsheetError as error:
        emit({
            "ok": False,
            "error": {
                "code": error.code,
                "message": error.message,
                "details": error.details,
                "blocked": error.blocked,
            },
        })
    except Exception as error:  # pragma: no cover - 最终诊断兜底
        emit({
            "ok": False,
            "error": {
                "code": "spreadsheet_worker_failed",
                "message": f"表格 worker 执行失败: {error}",
                "details": {"type": type(error).__name__},
            },
        })


def emit(value: dict[str, Any]) -> None:
    sys.stdout.write(json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=json_default))


def inspect_spreadsheet(workspace: Path, arguments: dict[str, Any]) -> dict[str, Any]:
    source_specs, legacy_single = normalize_inspect_sources(arguments)
    resolved_sources = resolve_inspect_sources(workspace, source_specs)
    sources: list[dict[str, Any]] = []
    sheets: list[dict[str, Any]] = []
    tables: list[dict[str, Any]] = []
    failed_sources: list[dict[str, Any]] = []
    warnings: list[str] = []

    for entry in resolved_sources:
        source_path = entry["path"]
        source_hash = hash_file(source_path)
        source_id = create_source_id(entry["relativePath"], source_hash)
        source = {
            "sourceId": source_id,
            "path": entry["relativePath"],
            "bytes": source_path.stat().st_size,
            "sha256": source_hash,
            "format": source_path.suffix.lower()[1:],
            "status": "ready",
        }
        try:
            source_sheets, source_tables, source_warnings = inspect_one_source(
                source_path,
                source_id,
                entry["sheets"],
            )
            if hash_file(source_path) != source_hash:
                raise SpreadsheetError(
                    "spreadsheet_source_changed",
                    "源表格在 inspect 过程中发生变化，请重新执行。",
                    {"sourceId": source_id, "path": entry["relativePath"]},
                    blocked=True,
                )
            for sheet in source_sheets:
                sheet.update({"sourceId": source_id, "sourcePath": entry["relativePath"]})
            for table in source_tables:
                table.update({"sourceId": source_id, "sourcePath": entry["relativePath"]})
            source.update({"sheetCount": len(source_sheets), "tableCount": len(source_tables)})
            if not source_tables:
                source["status"] = "no_tables"
                failed_sources.append({
                    "sourceId": source_id,
                    "path": entry["relativePath"],
                    "code": "spreadsheet_no_table_detected",
                    "message": "没有识别到包含表头和数据行的表格区域。",
                })
            sheets.extend(source_sheets)
            tables.extend(source_tables)
            warnings.extend(f"{entry['relativePath']}: {message}" for message in source_warnings)
        except SpreadsheetError as error:
            if legacy_single or error.code in {"spreadsheet_archive_unsafe", "spreadsheet_path_escape", "spreadsheet_source_changed"}:
                raise
            source["status"] = "failed"
            source.update({"sheetCount": 0, "tableCount": 0})
            failed_sources.append({
                "sourceId": source_id,
                "path": entry["relativePath"],
                "code": error.code,
                "message": error.message,
                "details": error.details,
            })
        sources.append(source)

    source_set_hash = create_source_set_hash(sources)
    duplicate_groups = detect_duplicate_groups(sources, tables)
    overlap_candidates = detect_overlap_candidates(tables)
    analysis_id = f"analysis-{source_set_hash[:12]}-{uuid.uuid4().hex[:8]}"
    inspection_status = "ready"
    if not tables:
        inspection_status = "blocked"
        warnings.append("没有识别到包含表头和数据行的表格区域。")
    elif failed_sources or duplicate_groups or overlap_candidates:
        inspection_status = "needs_review"
    elif any(table["needsSelection"] for table in tables) or any(sheet["tableCount"] > 1 for sheet in sheets):
        inspection_status = "needs_selection"

    manifest = {
        "schemaVersion": "agent-spreadsheet.analysis.v2",
        "analysisId": analysis_id,
        "createdAt": iso_now(),
        "sourceSetHash": source_set_hash,
        "sources": sources,
        "inspectionStatus": inspection_status,
        "sheets": sheets,
        "tables": tables,
        "duplicateGroups": duplicate_groups,
        "overlapCandidates": overlap_candidates,
        "failedSources": failed_sources,
        "warnings": warnings,
    }
    if len(sources) == 1:
        manifest["source"] = {key: sources[0][key] for key in ("path", "bytes", "sha256", "format")}
    storage_root = ensure_safe_directory(workspace, "temp")
    storage_root = ensure_safe_directory(storage_root, "spreadsheets")
    analysis_directory = storage_root / analysis_id
    analysis_directory.mkdir(exist_ok=False)
    manifest_path = analysis_directory / "manifest.json"
    atomic_write_json(manifest_path, manifest)
    public_tables = [public_table_summary(table) for table in tables[:200]]
    if len(tables) > len(public_tables):
        warnings.append("识别到的表格超过 200 个；当前只返回前 200 个，请通过 sheets 缩小检查范围。")
    return {
        "analysisId": analysis_id,
        "inspectionStatus": inspection_status,
        "sourceSetHash": source_set_hash,
        "source": manifest.get("source"),
        "sources": sources,
        "sheets": sheets,
        "tables": public_tables,
        "duplicateGroups": duplicate_groups,
        "overlapCandidates": overlap_candidates,
        "failedSources": failed_sources,
        "warnings": warnings,
        "manifestPath": relative_posix(workspace, manifest_path),
        "guidance": (
            "来源集合存在失败、重复或日期重叠；先明确来源取舍，再进行正式计算。"
            if inspection_status == "needs_review"
            else "存在多个或不确定的表格区域；计算时必须明确传入返回的 tableId。"
            if inspection_status == "needs_selection"
            else "计算时使用这里返回的 tableId，不要根据工作表名称猜测数据范围。"
        ),
    }


def normalize_inspect_sources(arguments: dict[str, Any]) -> tuple[list[dict[str, Any]], bool]:
    has_path = isinstance(arguments.get("path"), str) and bool(arguments["path"].strip())
    has_sources = arguments.get("sources") is not None
    if has_path == has_sources:
        raise SpreadsheetError("spreadsheet_sources_invalid", "path 与 sources 必须且只能提供一个。", blocked=True)
    if has_path:
        sheets = normalize_string_list(arguments.get("sheets"), "sheets", maximum=50)
        return [{"path": arguments["path"], "sheets": sheets}], True
    if arguments.get("sheets") is not None:
        raise SpreadsheetError("spreadsheet_sources_invalid", "使用 sources 时，sheets 必须写在对应来源对象内。", blocked=True)
    raw_sources = arguments.get("sources")
    if not isinstance(raw_sources, list) or not raw_sources or len(raw_sources) > MAX_SOURCES:
        raise SpreadsheetError("spreadsheet_sources_invalid", f"sources 必须是 1-{MAX_SOURCES} 项数组。", blocked=True)
    normalized = []
    for index, source in enumerate(raw_sources):
        if not isinstance(source, dict) or set(source) - {"path", "sheets"}:
            raise SpreadsheetError("spreadsheet_sources_invalid", f"sources[{index}] 只支持 path 和 sheets。", blocked=True)
        if not isinstance(source.get("path"), str) or not source["path"].strip():
            raise SpreadsheetError("spreadsheet_sources_invalid", f"sources[{index}].path 必须是非空字符串。", blocked=True)
        normalized.append({
            "path": source["path"],
            "sheets": normalize_string_list(source.get("sheets"), f"sources[{index}].sheets", maximum=50),
        })
    return normalized, False


def resolve_inspect_sources(workspace: Path, source_specs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    resolved = []
    seen_paths: set[str] = set()
    total_bytes = 0
    for source in source_specs:
        source_path = resolve_source_path(workspace, source["path"])
        canonical = os.path.normcase(str(source_path))
        if canonical in seen_paths:
            raise SpreadsheetError(
                "spreadsheet_sources_invalid",
                "同一个表格路径不能重复传入。",
                {"path": relative_posix(workspace, source_path)},
                blocked=True,
            )
        seen_paths.add(canonical)
        total_bytes += source_path.stat().st_size
        if total_bytes > MAX_TOTAL_SOURCE_BYTES:
            raise SpreadsheetError(
                "spreadsheet_sources_too_large",
                "表格来源总大小超过 1GB 安全上限。",
                {"bytes": total_bytes, "maxBytes": MAX_TOTAL_SOURCE_BYTES},
                blocked=True,
            )
        resolved.append({
            "path": source_path,
            "relativePath": relative_posix(workspace, source_path),
            "sheets": source["sheets"],
        })
    return resolved


def inspect_one_source(source_path: Path, source_id: str, requested_sheets: list[str]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    extension = source_path.suffix.lower()
    if extension in {".xlsx", ".xlsm"}:
        ensure_safe_archive(source_path)
        return inspect_excel(source_path, source_id, requested_sheets)
    return inspect_delimited(source_path, source_id, extension, requested_sheets)


def create_source_id(relative_path: str, source_hash: str) -> str:
    digest = hashlib.sha256(f"{relative_path}:{source_hash}".encode("utf8")).hexdigest()[:16]
    return f"source-{digest}"


def create_source_set_hash(sources: list[dict[str, Any]]) -> str:
    entries = sorted(f"{item['path']}:{item['sha256']}" for item in sources)
    return hashlib.sha256("\n".join(entries).encode("utf8")).hexdigest()


def detect_duplicate_groups(sources: list[dict[str, Any]], tables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: list[dict[str, Any]] = []
    by_file_hash: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for source in sources:
        by_file_hash[source["sha256"]].append(source)
    for source_hash, matches in by_file_hash.items():
        if len(matches) > 1:
            groups.append({
                "type": "identical_file",
                "sourceIds": [item["sourceId"] for item in matches],
                "paths": [item["path"] for item in matches],
                "sha256": source_hash,
            })
    by_table_hash: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for table in tables:
        by_table_hash[table["contentHash"]].append(table)
    for content_hash, matches in by_table_hash.items():
        source_ids = {item["sourceId"] for item in matches}
        if len(matches) > 1 and len(source_ids) > 1:
            groups.append({
                "type": "identical_table",
                "tableIds": [item["tableId"] for item in matches],
                "sourceIds": sorted(source_ids),
                "contentHash": content_hash,
            })
    return groups


def detect_overlap_candidates(tables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for left_index, left in enumerate(tables):
        for right in tables[left_index + 1:]:
            if left["sourceId"] == right["sourceId"] or left["schemaFingerprint"] != right["schemaFingerprint"]:
                continue
            left_ranges = {item["column"]: item for item in left.get("dateRanges") or []}
            right_ranges = {item["column"]: item for item in right.get("dateRanges") or []}
            for column in sorted(set(left_ranges) & set(right_ranges)):
                left_range = left_ranges[column]
                right_range = right_ranges[column]
                if left_range["minimum"] <= right_range["maximum"] and right_range["minimum"] <= left_range["maximum"]:
                    output.append({
                        "tableIds": [left["tableId"], right["tableId"]],
                        "sourceIds": [left["sourceId"], right["sourceId"]],
                        "dateColumn": column,
                        "ranges": [
                            [left_range["minimum"], left_range["maximum"]],
                            [right_range["minimum"], right_range["maximum"]],
                        ],
                    })
    return output


def inspect_excel(source_path: Path, source_hash: str, requested_sheets: list[str]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    keep_vba = source_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(source_path, read_only=False, data_only=False, keep_vba=keep_vba)
    cached_workbook = load_workbook(source_path, read_only=False, data_only=True, keep_vba=keep_vba)
    try:
        selected_names = requested_sheets or workbook.sheetnames
        missing = [name for name in selected_names if name not in workbook.sheetnames]
        if missing:
            raise SpreadsheetError("spreadsheet_sheet_not_found", "指定工作表不存在。", {"sheets": missing}, blocked=True)
        sheets: list[dict[str, Any]] = []
        tables: list[dict[str, Any]] = []
        warnings: list[str] = []
        non_empty_total = 0
        for sheet_name in selected_names:
            worksheet = workbook[sheet_name]
            cached_sheet = cached_workbook[sheet_name]
            ensure_cell_range_size(
                1,
                1,
                worksheet.max_row,
                worksheet.max_column,
                label=f"工作表 {sheet_name} 的已用区域",
            )
            explicit_ranges: list[tuple[int, int, int, int, str]] = []
            for table in worksheet.tables.values():
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                ensure_cell_range_size(
                    min_row,
                    min_col,
                    max_row,
                    max_col,
                    label=f"Excel Table {table.name}",
                )
                explicit_ranges.append((min_row, min_col, max_row, max_col, table.name))

            non_empty: set[tuple[int, int]] = set()
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None and cell.coordinate not in worksheet.merged_cells:
                        non_empty.add((cell.row, cell.column))
            for merged in worksheet.merged_cells.ranges:
                if worksheet.cell(merged.min_row, merged.min_col).value is not None:
                    non_empty.update(
                        (row, col)
                        for row in range(merged.min_row, merged.max_row + 1)
                        for col in range(merged.min_col, merged.max_col + 1)
                    )
            non_empty_total += len(non_empty)
            if non_empty_total > MAX_NON_EMPTY_CELLS:
                raise SpreadsheetError(
                    "spreadsheet_too_large",
                    "工作簿非空单元格超过安全上限。",
                    {"maxNonEmptyCells": MAX_NON_EMPTY_CELLS},
                    blocked=True,
                )

            sheet_tables: list[dict[str, Any]] = []
            for min_row, min_col, max_row, max_col, table_name in explicit_ranges:
                sheet_tables.append(build_excel_table_profile(
                    worksheet,
                    cached_sheet,
                    source_hash,
                    min_row,
                    min_col,
                    max_row,
                    max_col,
                    kind="excel_table",
                    declared_name=table_name,
                ))

            remaining = {
                (row, col)
                for row, col in non_empty
                if not any(
                    min_row <= row <= max_row and min_col <= col <= max_col
                    for min_row, min_col, max_row, max_col, _name in explicit_ranges
                )
            }
            for min_row, min_col, max_row, max_col in connected_ranges(remaining):
                if max_row <= min_row:
                    continue
                sheet_tables.append(build_excel_table_profile(
                    worksheet,
                    cached_sheet,
                    source_hash,
                    min_row,
                    min_col,
                    max_row,
                    max_col,
                    kind="detected_region",
                ))

            for table in sheet_tables:
                if len(sheet_tables) > 1:
                    table["needsSelection"] = True
                tables.append(table)
            sheets.append({
                "name": sheet_name,
                "state": worksheet.sheet_state,
                "maxRow": worksheet.max_row,
                "maxColumn": worksheet.max_column,
                "tableCount": len(sheet_tables),
                "mergedRanges": [str(item) for item in worksheet.merged_cells.ranges][:100],
            })
            if worksheet.sheet_state != "visible":
                warnings.append(f"工作表 {sheet_name} 处于 {worksheet.sheet_state} 状态。")
        return sheets, tables, warnings
    finally:
        workbook.close()
        cached_workbook.close()


def inspect_delimited(source_path: Path, source_hash: str, extension: str, requested_sheets: list[str]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    if requested_sheets and requested_sheets != ["data"]:
        raise SpreadsheetError("spreadsheet_sheet_not_found", "CSV/TSV 只有 data 工作表。", {"sheets": requested_sheets}, blocked=True)
    delimiter = "\t" if extension == ".tsv" else ","
    rows = read_delimited_rows(source_path, delimiter)
    if len(rows) < 2:
        return [{"name": "data", "state": "visible", "maxRow": len(rows), "maxColumn": len(rows[0]) if rows else 0, "tableCount": 0, "mergedRanges": []}], [], []
    width = max(len(row) for row in rows)
    if width > MAX_TABLE_COLUMNS:
        raise SpreadsheetError("spreadsheet_too_wide", "表格列数超过安全上限。", {"maxColumns": MAX_TABLE_COLUMNS}, blocked=True)
    normalized = [row + [None] * (width - len(row)) for row in rows]
    table = build_table_profile(
        sheet_name="data",
        source_hash=source_hash,
        values=normalized,
        formula_values=None,
        min_row=1,
        min_col=1,
        max_row=len(normalized),
        max_col=width,
        kind="delimited",
    )
    return [{"name": "data", "state": "visible", "maxRow": len(rows), "maxColumn": width, "tableCount": 1, "mergedRanges": []}], [table], []


def build_excel_table_profile(worksheet: Any, cached_sheet: Any, source_hash: str, min_row: int, min_col: int, max_row: int, max_col: int, kind: str, declared_name: str | None = None) -> dict[str, Any]:
    ensure_cell_range_size(min_row, min_col, max_row, max_col, label=declared_name or "候选表格区域")
    if max_col - min_col + 1 > MAX_TABLE_COLUMNS:
        raise SpreadsheetError("spreadsheet_too_wide", "表格列数超过安全上限。", {"maxColumns": MAX_TABLE_COLUMNS}, blocked=True)
    formula_values = [
        [worksheet.cell(row, col).value for col in range(min_col, max_col + 1)]
        for row in range(min_row, max_row + 1)
    ]
    cached_values = [
        [cached_sheet.cell(row, col).value for col in range(min_col, max_col + 1)]
        for row in range(min_row, max_row + 1)
    ]
    return build_table_profile(
        sheet_name=worksheet.title,
        source_hash=source_hash,
        values=cached_values,
        formula_values=formula_values,
        min_row=min_row,
        min_col=min_col,
        max_row=max_row,
        max_col=max_col,
        kind=kind,
        declared_name=declared_name,
        merged_ranges=[
            str(item)
            for item in worksheet.merged_cells.ranges
            if ranges_overlap(min_row, min_col, max_row, max_col, item.min_row, item.min_col, item.max_row, item.max_col)
        ],
    )


def build_table_profile(sheet_name: str, source_hash: str, values: list[list[Any]], formula_values: list[list[Any]] | None, min_row: int, min_col: int, max_row: int, max_col: int, kind: str, declared_name: str | None = None, merged_ranges: list[str] | None = None) -> dict[str, Any]:
    raw_headers = values[0] if values else []
    headers, header_warnings = normalize_headers(raw_headers)
    data_rows = values[1:]
    formula_rows = formula_values[1:] if formula_values else None
    profiles: list[dict[str, Any]] = []
    formula_columns: list[str] = []
    for index, header in enumerate(headers):
        column_values = [row[index] if index < len(row) else None for row in data_rows]
        formula_count = 0
        cached_count = 0
        if formula_rows is not None:
            for row_index, row in enumerate(formula_rows):
                value = row[index] if index < len(row) else None
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                    cached_value = data_rows[row_index][index] if row_index < len(data_rows) and index < len(data_rows[row_index]) else None
                    if cached_value is not None:
                        cached_count += 1
        if formula_count:
            formula_columns.append(header)
        profiles.append(profile_column(header, column_values, formula_count, cached_count))

    cell_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    table_id = f"table-{hashlib.sha256(f'{source_hash}:{sheet_name}:{cell_range}'.encode('utf8')).hexdigest()[:16]}"
    schema_fingerprint = hashlib.sha256(json.dumps(
        [{"name": item["name"], "type": item["inferredType"]} for item in profiles],
        ensure_ascii=False,
        sort_keys=True,
    ).encode("utf8")).hexdigest()
    content_hash = hash_table_values(values)
    date_ranges = [
        {"column": item["name"], **item["dateRange"]}
        for item in profiles
        if item.get("dateRange")
    ]
    structure_ambiguous = bool(header_warnings or merged_ranges)
    summary_rows = detect_summary_rows(data_rows, min_row + 1)
    return {
        "tableId": table_id,
        "sheet": sheet_name,
        "range": cell_range,
        "kind": kind,
        "declaredName": declared_name,
        "startRow": min_row,
        "startColumn": min_col,
        "endRow": max_row,
        "endColumn": max_col,
        "rowCount": max(0, len(data_rows)),
        "columnCount": len(headers),
        "schemaFingerprint": schema_fingerprint,
        "contentHash": content_hash,
        "dateRanges": date_ranges,
        "headers": headers,
        "headerWarnings": header_warnings,
        "columns": profiles,
        "formulaColumns": formula_columns,
        "summaryRows": summary_rows,
        "mergedRanges": merged_ranges or [],
        "structureAmbiguous": structure_ambiguous,
        "needsSelection": structure_ambiguous,
        "sampleRows": [serialize_row(headers, row) for row in data_rows[:5]],
    }


def profile_column(name: str, values: list[Any], formula_count: int, cached_count: int) -> dict[str, Any]:
    non_null = [value for value in values if value not in (None, "")]
    type_counts = Counter(value_type(value) for value in non_null)
    serial_values = [canonical_scalar(value) for value in non_null]
    duplicate_count = sum(count - 1 for count in Counter(serial_values).values() if count > 1)
    ambiguous_numeric = sum(1 for value in non_null if isinstance(value, str) and looks_ambiguous_numeric(value))
    profile = {
        "name": name,
        "inferredType": infer_type(type_counts),
        "nonNullCount": len(non_null),
        "nullCount": len(values) - len(non_null),
        "duplicateValueCount": duplicate_count,
        "typeCounts": dict(type_counts),
        "ambiguousNumericCount": ambiguous_numeric,
        "formulaCount": formula_count,
        "cachedFormulaValueCount": cached_count,
        "formulaStatus": "cached_unverified" if cached_count else "formula_backed" if formula_count else "none",
    }
    date_values = [value for value in non_null if isinstance(value, (date, datetime))]
    if date_values and len(date_values) == len(non_null):
        profile["dateRange"] = {
            "minimum": serialize_value(min(date_values)),
            "maximum": serialize_value(max(date_values)),
        }
    elif non_null and all(isinstance(value, str) for value in non_null):
        parsed_dates = [parse_iso_date_candidate(value) for value in non_null]
        if all(value is not None for value in parsed_dates):
            profile["dateRange"] = {
                "minimum": min(parsed_dates),
                "maximum": max(parsed_dates),
            }
    return profile


def hash_table_values(values: list[list[Any]]) -> str:
    digest = hashlib.sha256()
    for row in values:
        digest.update(json.dumps(
            row,
            ensure_ascii=False,
            separators=(",", ":"),
            default=json_default,
        ).encode("utf8"))
        digest.update(b"\n")
    return digest.hexdigest()


def parse_iso_date_candidate(value: str) -> str | None:
    text = value.strip()
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}(?:[T ][0-9:.+-Z]+)?", text):
        return None
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00")) if len(text) > 10 else date.fromisoformat(text)
        return parsed.isoformat()
    except ValueError:
        return None


def detect_summary_rows(rows: list[list[Any]], first_source_row: int) -> list[dict[str, Any]]:
    labels = {"total", "grand total", "subtotal", "合计", "总计", "小计"}
    output = []
    for index, row in enumerate(rows):
        first_value = next((value for value in row if value not in (None, "")), None)
        if not isinstance(first_value, str) or first_value.strip().lower() not in labels:
            continue
        if not any(isinstance(value, (int, float, Decimal)) and not isinstance(value, bool) for value in row[1:]):
            continue
        output.append({
            "rowIndex": index,
            "sourceRow": first_source_row + index,
            "label": first_value.strip(),
        })
    return output


def apply_summary_row_policy(metadata: dict[str, Any], rows: list[dict[str, Any]], policy: Any) -> list[dict[str, Any]]:
    summary_rows = metadata.get("summaryRows") or []
    if not summary_rows:
        if policy not in (None, "include", "exclude"):
            raise SpreadsheetError("spreadsheet_summary_row_policy_invalid", "summaryRowPolicy 只支持 include 或 exclude。", blocked=True)
        return rows
    if policy is None:
        raise SpreadsheetError(
            "spreadsheet_summary_rows_ambiguous",
            "检测到合计/总计行；必须显式选择 include 或 exclude，避免金额重复计算。",
            {"tableId": metadata["tableId"], "summaryRows": summary_rows},
            blocked=True,
        )
    if policy not in {"include", "exclude"}:
        raise SpreadsheetError("spreadsheet_summary_row_policy_invalid", "summaryRowPolicy 只支持 include 或 exclude。", blocked=True)
    if policy == "include":
        return rows
    excluded = {item["rowIndex"] for item in summary_rows}
    return [row for index, row in enumerate(rows) if index not in excluded]


def compute_spreadsheet(workspace: Path, arguments: dict[str, Any]) -> dict[str, Any]:
    analysis_id = require_id(arguments.get("analysisId"), "analysisId", "analysis")
    manifest, analysis_directory, source_paths = load_analysis(workspace, analysis_id)
    ensure_sources_unchanged(manifest, source_paths)
    queries = arguments.get("queries")
    if not isinstance(queries, list) or not queries or len(queries) > MAX_QUERIES:
        raise SpreadsheetError("spreadsheet_queries_invalid", f"queries 必须是 1-{MAX_QUERIES} 项数组。", blocked=True)
    query_ids: set[str] = set()
    results: list[dict[str, Any]] = []
    warnings: list[str] = []
    context = WorkbookContext(source_paths, manifest)
    source_decisions = normalize_source_decisions(arguments.get("sourceDecisions"), manifest)
    try:
        computed: list[dict[str, Any]] = []
        for query in queries:
            if not isinstance(query, dict):
                raise SpreadsheetError("spreadsheet_query_invalid", "每个 query 必须是对象。", blocked=True)
            query_id = require_name(query.get("id"), "query.id")
            if query_id in query_ids:
                raise SpreadsheetError("spreadsheet_query_invalid", f"query.id 重复: {query_id}", blocked=True)
            query_ids.add(query_id)
            computed.append(execute_query(context, query))

        ensure_sources_unchanged(manifest, source_paths)
        results_directory = ensure_safe_directory(analysis_directory, "results")
        for query, result in zip(queries, computed):
            result_id = f"result-{uuid.uuid4().hex}"
            data_payload = {
                "schemaVersion": "agent-spreadsheet.data.v1",
                "analysisId": analysis_id,
                "resultId": result_id,
                "queryId": query["id"],
                "columns": result["columns"],
                "rows": result["rows"],
                "notComputable": result["notComputable"],
                "notComputableCount": result["notComputableCount"],
                "notComputableTruncated": result["notComputableTruncated"],
            }
            data_path = results_directory / f"{result_id}.data.json"
            csv_path = results_directory / f"{result_id}.csv"
            atomic_write_json(data_path, data_payload)
            write_result_csv(csv_path, result["columns"], result["rows"])
            data_hash = hash_file(data_path)
            csv_hash = hash_file(csv_path)
            lineage = {
                "sourceSetHash": manifest["sourceSetHash"],
                "sources": result_source_lineage(manifest, result["tableIds"]),
                "tables": [table_lineage(manifest, table_id) for table_id in result["tableIds"]],
                "sourceDecisions": source_decisions,
                "filters": query.get("filters") or [],
                "from": query.get("from"),
                "joins": query.get("joins") or [],
                "groupBy": query.get("groupBy") or [],
                "measures": query.get("measures") or [],
                "derivedMetrics": query.get("derivedMetrics") or [],
                "inputRowCount": result["inputRowCount"],
                "joinedRowCount": result["joinedRowCount"],
                "filteredRowCount": result["filteredRowCount"],
                "outputRowCount": result["rowCount"],
            }
            result_manifest = {
                "schemaVersion": "agent-spreadsheet.result.v1",
                "analysisId": analysis_id,
                "resultId": result_id,
                "queryId": query["id"],
                "createdAt": iso_now(),
                "sourceSetHash": manifest["sourceSetHash"],
                "tableIds": result["tableIds"],
                "sourceDecisions": source_decisions,
                "query": query,
                "columns": result["columns"],
                "rowCount": result["rowCount"],
                "returnedRows": len(result["rows"]),
                "truncated": result["truncated"],
                "dataFile": f"results/{result_id}.data.json",
                "dataHash": data_hash,
                "csvFile": f"results/{result_id}.csv",
                "csvHash": csv_hash,
                "notComputable": result["notComputable"][:MAX_PUBLIC_NOT_COMPUTABLE_SAMPLES],
                "notComputableCount": result["notComputableCount"],
                "notComputableTruncated": result["notComputableTruncated"],
                "notComputableSamplesTruncated": result["notComputableCount"] > MAX_PUBLIC_NOT_COMPUTABLE_SAMPLES,
                "lineage": lineage,
            }
            if len(manifest["sources"]) == 1:
                result_manifest["sourceHash"] = manifest["sources"][0]["sha256"]
            result_manifest_path = results_directory / f"{result_id}.manifest.json"
            atomic_write_json(result_manifest_path, result_manifest)
            summary = {
                "queryId": query["id"],
                "resultId": result_id,
                "rowCount": result["rowCount"],
                "returnedRows": len(result["rows"]),
                "truncated": result["truncated"],
                "columns": result["columns"][:100],
                "columnsTruncated": len(result["columns"]) > 100,
                "preview": compact_result_preview(result["rows"]),
                "notComputable": result["notComputable"][:MAX_PUBLIC_NOT_COMPUTABLE_SAMPLES],
                "notComputableCount": result["notComputableCount"],
                "notComputableTruncated": result["notComputableTruncated"],
                "notComputableSamplesTruncated": result["notComputableCount"] > MAX_PUBLIC_NOT_COMPUTABLE_SAMPLES,
                "lineage": {
                    "sourceSetHash": lineage["sourceSetHash"],
                    "sourceIds": [source["sourceId"] for source in lineage["sources"]],
                    "tables": [
                        {
                            "tableId": table["tableId"],
                            "sourceId": table.get("sourceId"),
                            "sheet": table.get("sheet"),
                            "range": table.get("range"),
                        }
                        for table in lineage["tables"]
                    ],
                    "inputRowCount": lineage["inputRowCount"],
                    "joinedRowCount": lineage["joinedRowCount"],
                    "filteredRowCount": lineage["filteredRowCount"],
                    "outputRowCount": lineage["outputRowCount"],
                },
                "dataRef": {
                    "schemaVersion": "agent-spreadsheet.data-ref.v1",
                    "analysisId": analysis_id,
                    "resultId": result_id,
                },
                "manifestPath": relative_posix(workspace, result_manifest_path),
                "csvPath": relative_posix(workspace, csv_path),
            }
            results.append(summary)
            if result["truncated"]:
                warnings.append(
                    f"查询 {query['id']} 显式限制为前 {result['returnedLimit']} 行；"
                    "该 canonical result 不能用于证明完整集合或通过质量门。"
                )
        return {"analysisId": analysis_id, "results": results, "warnings": warnings}
    finally:
        context.close()


def normalize_source_decisions(value: Any, manifest: dict[str, Any]) -> list[dict[str, Any]]:
    if value is None:
        return []
    if not isinstance(value, list) or len(value) > MAX_SOURCES:
        raise SpreadsheetError("spreadsheet_source_decisions_invalid", f"sourceDecisions 必须是最多 {MAX_SOURCES} 项数组。", blocked=True)
    known = {source["sourceId"] for source in manifest["sources"]}
    known_paths = {source["path"]: source["sourceId"] for source in manifest["sources"]}
    allowed_reasons = {
        "exact_duplicate",
        "out_of_scope",
        "unsupported_period",
        "corrupt_source",
        "superseded",
        "user_excluded",
        "other",
    }
    seen: set[str] = set()
    output = []
    for index, decision in enumerate(value):
        if not isinstance(decision, dict) or set(decision) - {"sourceId", "action", "reasonCode", "reason"}:
            raise SpreadsheetError("spreadsheet_source_decisions_invalid", f"sourceDecisions[{index}] 字段无效。", blocked=True)
        source_id = require_name(decision.get("sourceId"), f"sourceDecisions[{index}].sourceId")
        # 模型偶尔会复制 inspect 返回的唯一来源 path。路径在同一 analysis 内唯一，
        # 因此可以安全归一化为真正 sourceId，持久化结果仍只保存规范 ID。
        source_id = known_paths.get(source_id, source_id)
        if source_id not in known or source_id in seen:
            raise SpreadsheetError("spreadsheet_source_decisions_invalid", f"sourceId 不存在或重复: {source_id}", blocked=True)
        action = decision.get("action")
        if action not in {"include", "exclude"}:
            raise SpreadsheetError("spreadsheet_source_decisions_invalid", "sourceDecisions.action 只支持 include/exclude。", blocked=True)
        normalized = {"sourceId": source_id, "action": action}
        if action == "exclude":
            reason_code = decision.get("reasonCode")
            reason = decision.get("reason")
            if reason_code not in allowed_reasons or not isinstance(reason, str) or not reason.strip():
                raise SpreadsheetError(
                    "spreadsheet_source_decisions_invalid",
                    "exclude 必须提供合法 reasonCode 和非空 reason。",
                    {"sourceId": source_id},
                    blocked=True,
                )
            normalized.update({"reasonCode": reason_code, "reason": reason.strip()[:500]})
        seen.add(source_id)
        output.append(normalized)
    return output


class WorkbookContext:
    """在一次 worker 调用中按需复用多个来源的工作簿读取器。"""

    def __init__(self, source_paths: dict[str, Path], manifest: dict[str, Any]):
        self.source_paths = source_paths
        self.manifest = manifest
        self.workbooks: dict[str, Any] = {}
        self.table_cache: dict[str, list[dict[str, Any]]] = {}

    def close(self) -> None:
        for workbook in self.workbooks.values():
            workbook.close()
        self.workbooks.clear()

    def table(self, table_id: str) -> tuple[dict[str, Any], list[dict[str, Any]]]:
        metadata = next((item for item in self.manifest["tables"] if item["tableId"] == table_id), None)
        if metadata is None:
            raise SpreadsheetError("spreadsheet_table_not_found", f"找不到 tableId: {table_id}", blocked=True)
        if metadata.get("structureAmbiguous") or metadata.get("headerWarnings") or metadata.get("mergedRanges"):
            raise SpreadsheetError(
                "spreadsheet_table_ambiguous",
                f"表格 {metadata['tableId']} 存在合并单元格或表头歧义，不能进行确定性计算。",
                {
                    "headerWarnings": metadata.get("headerWarnings") or [],
                    "mergedRanges": metadata.get("mergedRanges") or [],
                },
                blocked=True,
            )
        if table_id not in self.table_cache:
            self.table_cache[table_id] = self._read_table(metadata)
        return metadata, [dict(row) for row in self.table_cache[table_id]]

    def _read_table(self, metadata: dict[str, Any]) -> list[dict[str, Any]]:
        source_id = metadata["sourceId"]
        source_path = self.source_paths[source_id]
        extension = source_path.suffix.lower()
        headers = metadata["headers"]
        if len(set(headers)) != len(headers):
            raise SpreadsheetError(
                "spreadsheet_header_ambiguous",
                f"表格 {metadata['tableId']} 的表头不适合确定性计算。",
                {"warnings": metadata.get("headerWarnings")},
                blocked=True,
            )
        rows: list[dict[str, Any]] = []
        if extension in {".xlsx", ".xlsm"}:
            if source_id not in self.workbooks:
                self.workbooks[source_id] = load_workbook(
                    source_path,
                    read_only=True,
                    data_only=True,
                    keep_vba=extension == ".xlsm",
                )
            worksheet = self.workbooks[source_id][metadata["sheet"]]
            for values in worksheet.iter_rows(
                min_row=metadata["startRow"] + 1,
                max_row=metadata["endRow"],
                min_col=metadata["startColumn"],
                max_col=metadata["endColumn"],
                values_only=True,
            ):
                rows.append({header: value for header, value in zip(headers, values)})
        else:
            delimiter = "\t" if extension == ".tsv" else ","
            raw = read_delimited_rows(source_path, delimiter)
            for values in raw[1:]:
                padded = values + [None] * (len(headers) - len(values))
                rows.append({header: value for header, value in zip(headers, padded)})
        return rows


def execute_query(context: WorkbookContext, query: dict[str, Any]) -> dict[str, Any]:
    table_id = query_default_table_id(query)
    parsers = normalize_parsers(query.get("columnParsers"), table_id)
    metadata, rows, table_ids = load_query_input(context, query, parsers)
    join_prefixes = normalized_join_prefixes(query.get("joins") or [])
    source_columns = referenced_columns(query, metadata["tableId"], join_prefixes, set(metadata.get("headers") or []))
    assert_columns_declared(metadata, source_columns)
    assert_authoritative_columns(metadata, source_columns)
    input_row_count = len(rows)
    available_fields = set(metadata.get("headers") or [])

    joins = query.get("joins") or []
    if not isinstance(joins, list) or len(joins) > 5:
        raise SpreadsheetError("spreadsheet_join_invalid", "joins 必须是最多 5 项的数组。", blocked=True)
    requested_fields = (
        query_input_fields(query)
        | sort_fields(query.get("sort") or [])
        | join_left_fields(joins)
    )
    for index, join in enumerate(joins):
        rows, joined_id, joined_fields = apply_join(
            context,
            rows,
            join,
            parsers,
            index,
            requested_fields,
            available_fields,
        )
        table_ids.append(joined_id)
        available_fields.update(joined_fields)

    unused_parser_tables = sorted(set(parsers) - set(table_ids))
    if unused_parser_tables:
        raise SpreadsheetError(
            "spreadsheet_parser_invalid",
            "columnParsers 引用了本次查询未使用的 tableId。",
            {"tableIds": unused_parser_tables},
            blocked=True,
        )

    missing_fields = sorted(query_input_fields(query) - available_fields)
    if missing_fields:
        joined_fields = sorted(
            field for field in available_fields
            if any(field.startswith(f"{prefix}.") for prefix in join_prefixes)
        )
        raise SpreadsheetError(
            "spreadsheet_column_not_found",
            (
                "查询引用了不存在的字段；联接右表非键字段必须写成 <prefix>.<column>。"
                if joins
                else "查询引用了不存在的字段。"
            ),
            {"columns": missing_fields, "availableJoinFields": joined_fields[:100]},
            blocked=True,
        )

    joined_row_count = len(rows)
    rows = apply_filters(rows, query.get("filters") or [])
    filtered_row_count = len(rows)
    columns = normalize_string_list(query.get("columns"), "query.columns", maximum=MAX_TABLE_COLUMNS)
    measures = query.get("measures") or []
    group_by = normalize_string_list(query.get("groupBy"), "query.groupBy", maximum=64)
    derived = query.get("derivedMetrics") or []
    if columns and (measures or group_by or derived):
        raise SpreadsheetError("spreadsheet_query_invalid", "columns 选择模式不能同时使用 groupBy/measures/derivedMetrics。", blocked=True)
    if not columns and not measures:
        raise SpreadsheetError("spreadsheet_query_invalid", "query 必须提供 columns 或 measures。", blocked=True)

    not_computable: list[dict[str, Any]] = []
    not_computable_count = 0
    if columns:
        ensure_columns(rows, columns)
        output_rows = [{column: row.get(column) for column in columns} for row in rows]
    else:
        output_rows = aggregate_rows(rows, group_by, measures, query.get("nullPolicy", "fail"), not_computable)
        aggregate_fields = set(group_by) | {
            measure["id"]
            for measure in measures
            if isinstance(measure, dict) and isinstance(measure.get("id"), str)
        }
        not_computable_count = apply_derived_metrics(output_rows, derived, not_computable, aggregate_fields)

    expected_columns = columns or [
        *group_by,
        *[measure["id"] for measure in measures if isinstance(measure, dict) and measure.get("id")],
        *[metric["id"] for metric in derived if isinstance(metric, dict) and metric.get("id")],
    ]
    output_rows = sort_rows(output_rows, query.get("sort") or [], set(expected_columns))
    row_count = len(output_rows)
    has_explicit_limit = "limit" in query
    requested_limit = query.get("limit", MAX_QUERY_ROWS)
    if not isinstance(requested_limit, int) or requested_limit < 1 or requested_limit > MAX_QUERY_ROWS:
        raise SpreadsheetError("spreadsheet_query_invalid", f"limit 必须是 1-{MAX_QUERY_ROWS} 的整数。", blocked=True)
    if row_count > MAX_QUERY_ROWS and not has_explicit_limit:
        raise SpreadsheetError(
            "spreadsheet_result_too_large",
            "查询结果超过安全持久化上限；请先筛选、聚合，或显式声明 limit 形成受限结果。",
            {"rowCount": row_count, "maxRows": MAX_QUERY_ROWS},
            blocked=True,
        )
    output_columns = infer_output_columns(output_rows, expected_columns)
    output_rows = [
        {name: serialize_value(value) for name, value in row.items()}
        for row in output_rows[:requested_limit]
    ]
    return {
        "tableIds": table_ids,
        "rows": output_rows,
        "columns": output_columns,
        "rowCount": row_count,
        "truncated": row_count > requested_limit,
        "returnedLimit": requested_limit,
        "notComputable": not_computable,
        "notComputableCount": not_computable_count,
        "notComputableTruncated": not_computable_count > len(not_computable),
        "inputRowCount": input_row_count,
        "joinedRowCount": joined_row_count,
        "filteredRowCount": filtered_row_count,
    }


def query_default_table_id(query: dict[str, Any]) -> str | None:
    source = query.get("from")
    legacy_table_id = query.get("tableId")
    if source is None:
        return require_name(legacy_table_id, "query.tableId")
    if legacy_table_id is not None:
        raise SpreadsheetError("spreadsheet_query_invalid", "query.tableId 与 query.from 不能同时提供。", blocked=True)
    if query.get("summaryRowPolicy") is not None:
        raise SpreadsheetError(
            "spreadsheet_query_invalid",
            "使用 query.from 时，summaryRowPolicy 必须写在 from 或 union.tables 条目中。",
            blocked=True,
        )
    if not isinstance(source, dict):
        raise SpreadsheetError("spreadsheet_query_invalid", "query.from 必须是对象。", blocked=True)
    source_type = source.get("type")
    if source_type == "table":
        if set(source) - {"type", "tableId", "summaryRowPolicy"}:
            raise SpreadsheetError("spreadsheet_query_invalid", "from.type=table 包含不支持字段。", blocked=True)
        return require_name(source.get("tableId"), "query.from.tableId")
    if source_type == "union":
        return None
    raise SpreadsheetError("spreadsheet_query_invalid", "query.from.type 只支持 table 或 union。", blocked=True)


def load_query_input(context: WorkbookContext, query: dict[str, Any], parsers: dict[str, dict[str, dict[str, Any]]]) -> tuple[dict[str, Any], list[dict[str, Any]], list[str]]:
    source = query.get("from")
    if source is None:
        table_id = require_name(query.get("tableId"), "query.tableId")
        metadata, rows = context.table(table_id)
        rows = apply_summary_row_policy(metadata, rows, query.get("summaryRowPolicy"))
        assert_columns_declared(metadata, set(parsers.get(table_id, {})))
        rows = apply_parsers(rows, parsers.get(table_id, {}))
        return metadata, rows, [table_id]
    if source["type"] == "table":
        table_id = require_name(source.get("tableId"), "query.from.tableId")
        metadata, rows = context.table(table_id)
        rows = apply_summary_row_policy(metadata, rows, source.get("summaryRowPolicy"))
        assert_columns_declared(metadata, set(parsers.get(table_id, {})))
        rows = apply_parsers(rows, parsers.get(table_id, {}))
        return metadata, rows, [table_id]
    return load_union_input(context, source, parsers)


def load_union_input(context: WorkbookContext, source: dict[str, Any], parsers: dict[str, dict[str, dict[str, Any]]]) -> tuple[dict[str, Any], list[dict[str, Any]], list[str]]:
    if set(source) - {"type", "tables"}:
        raise SpreadsheetError("spreadsheet_union_invalid", "from.type=union 包含不支持字段。", blocked=True)
    entries = source.get("tables")
    if not isinstance(entries, list) or len(entries) < 2 or len(entries) > MAX_SOURCES:
        raise SpreadsheetError("spreadsheet_union_invalid", f"union.tables 必须是 2-{MAX_SOURCES} 项数组。", blocked=True)
    table_ids: list[str] = []
    combined_rows: list[dict[str, Any]] = []
    canonical_headers: list[str] | None = None
    canonical_types: dict[str, str] | None = None
    formula_columns: set[str] = set()
    source_metadata: list[dict[str, Any]] = []
    for index, entry in enumerate(entries):
        if not isinstance(entry, dict) or set(entry) - {"tableId", "columnMap", "summaryRowPolicy"}:
            raise SpreadsheetError("spreadsheet_union_invalid", f"union.tables[{index}] 字段无效。", blocked=True)
        table_id = require_name(entry.get("tableId"), f"union.tables[{index}].tableId")
        if table_id in table_ids:
            raise SpreadsheetError("spreadsheet_union_invalid", f"union tableId 重复: {table_id}", blocked=True)
        metadata, rows = context.table(table_id)
        rows = apply_summary_row_policy(metadata, rows, entry.get("summaryRowPolicy"))
        assert_columns_declared(metadata, set(parsers.get(table_id, {})))
        rows = apply_parsers(rows, parsers.get(table_id, {}))
        column_map = normalize_union_column_map(entry.get("columnMap"), metadata, index)
        selected_headers = list(column_map)
        mapped_headers = list(column_map.values())
        if len(set(mapped_headers)) != len(mapped_headers):
            raise SpreadsheetError("spreadsheet_union_invalid", f"union.tables[{index}].columnMap 目标字段重复。", blocked=True)
        mapped_types = {
            column_map[item["name"]]: parser_output_type(parsers.get(table_id, {}).get(item["name"]), item["inferredType"])
            for item in metadata["columns"]
            if item["name"] in column_map
        }
        if canonical_headers is None:
            canonical_headers = mapped_headers
            canonical_types = mapped_types
        elif set(mapped_headers) != set(canonical_headers):
            raise SpreadsheetError(
                "spreadsheet_union_schema_mismatch",
                "union 各表映射后的字段集合不一致。",
                {"tableId": table_id, "expected": canonical_headers, "actual": mapped_headers},
                blocked=True,
            )
        else:
            type_mismatches = [
                {"column": name, "expected": canonical_types[name], "actual": mapped_types[name]}
                for name in canonical_headers
                if not compatible_union_types(canonical_types[name], mapped_types[name])
            ]
            if type_mismatches:
                raise SpreadsheetError(
                    "spreadsheet_union_schema_mismatch",
                    "union 各表映射后的字段类型不一致；请使用 columnParsers 显式统一。",
                    {"tableId": table_id, "columns": type_mismatches},
                    blocked=True,
                )
        for row in rows:
            combined_rows.append({target: row.get(source_name) for source_name, target in column_map.items()})
        formula_columns.update(
            column_map[name]
            for name in metadata.get("formulaColumns") or []
            if name in column_map
        )
        table_ids.append(table_id)
        source_metadata.append(metadata)
    synthetic_id = f"union-{hashlib.sha256(':'.join(table_ids).encode('utf8')).hexdigest()[:16]}"
    metadata = {
        "tableId": synthetic_id,
        "headers": canonical_headers or [],
        "columns": [
            {"name": name, "inferredType": (canonical_types or {}).get(name, "empty")}
            for name in canonical_headers or []
        ],
        "formulaColumns": sorted(formula_columns),
        "summaryRows": [],
        "sourceTables": [item["tableId"] for item in source_metadata],
    }
    return metadata, combined_rows, table_ids


def normalize_union_column_map(value: Any, metadata: dict[str, Any], index: int) -> dict[str, str]:
    headers = metadata.get("headers") or []
    if value is None:
        return {name: name for name in headers}
    if not isinstance(value, dict) or not value:
        raise SpreadsheetError("spreadsheet_union_invalid", f"union.tables[{index}].columnMap 必须是非空对象。", blocked=True)
    output: dict[str, str] = {}
    for source_name, target_name in value.items():
        if not isinstance(source_name, str) or source_name not in headers:
            raise SpreadsheetError("spreadsheet_column_not_found", f"union 来源字段不存在: {source_name}", {"tableId": metadata["tableId"]}, blocked=True)
        output[source_name] = require_name(target_name, f"union.tables[{index}].columnMap.{source_name}")
    return output


def parser_output_type(parser: dict[str, Any] | None, inferred_type: str) -> str:
    return parser["type"] if parser else inferred_type


def compatible_union_types(left: str, right: str) -> bool:
    if left == right:
        return True
    return {left, right} <= {"integer", "decimal", "number"}


def normalize_parsers(value: Any, default_table_id: str) -> dict[str, dict[str, dict[str, Any]]]:
    if value is None:
        return {}
    if not isinstance(value, list) or len(value) > 100:
        raise SpreadsheetError("spreadsheet_parser_invalid", "columnParsers 必须是最多 100 项数组。", blocked=True)
    output: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    for parser in value:
        if not isinstance(parser, dict):
            raise SpreadsheetError("spreadsheet_parser_invalid", "columnParser 必须是对象。", blocked=True)
        table_id = parser.get("tableId") or default_table_id
        if table_id is None:
            raise SpreadsheetError(
                "spreadsheet_parser_invalid",
                "union 查询中的每个 columnParser 都必须明确提供 tableId。",
                blocked=True,
            )
        column = require_name(parser.get("column"), "columnParser.column")
        parser_type = parser.get("type")
        if parser_type not in {"decimal", "integer", "string", "date"}:
            raise SpreadsheetError("spreadsheet_parser_invalid", f"不支持的列解析类型: {parser_type}", blocked=True)
        decimal_separator = parser.get("decimalSeparator", ".")
        thousands_separator = parser.get("thousandsSeparator")
        if decimal_separator not in {".", ","}:
            raise SpreadsheetError("spreadsheet_parser_invalid", "decimalSeparator 只支持 . 或 ,。", blocked=True)
        if thousands_separator is not None:
            if thousands_separator not in {".", ",", " ", "'"}:
                raise SpreadsheetError("spreadsheet_parser_invalid", "thousandsSeparator 只支持 .、,、空格或单引号。", blocked=True)
            if thousands_separator == decimal_separator:
                raise SpreadsheetError("spreadsheet_parser_invalid", "小数分隔符和千位分隔符不能相同。", blocked=True)
        currency_symbols = parser.get("currencySymbols") or []
        if not isinstance(currency_symbols, list) or len(currency_symbols) > 10 or any(
            not isinstance(symbol, str) or not symbol or len(symbol) > 10
            for symbol in currency_symbols
        ):
            raise SpreadsheetError("spreadsheet_parser_invalid", "currencySymbols 必须是最多 10 项的短字符串数组。", blocked=True)
        if column in output[table_id]:
            raise SpreadsheetError(
                "spreadsheet_parser_invalid",
                f"columnParser 重复声明: {table_id}/{column}",
                blocked=True,
            )
        output[table_id][column] = parser
    return output


def apply_parsers(rows: list[dict[str, Any]], parsers: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    if not parsers:
        return rows
    for column in parsers:
        ensure_columns(rows, [column])
    output = []
    for row in rows:
        updated = dict(row)
        for column, parser in parsers.items():
            updated[column] = parse_typed_value(row.get(column), parser)
        output.append(updated)
    return output


def apply_join(context: WorkbookContext, left_rows: list[dict[str, Any]], join: Any, parsers: dict[str, dict[str, dict[str, Any]]], index: int, requested_fields: set[str], available_fields: set[str]) -> tuple[list[dict[str, Any]], str, set[str]]:
    if not isinstance(join, dict):
        raise SpreadsheetError("spreadsheet_join_invalid", "join 必须是对象。", blocked=True)
    right_table_id = require_name(join.get("tableId"), "join.tableId")
    right_metadata, right_rows = context.table(right_table_id)
    right_rows = apply_summary_row_policy(right_metadata, right_rows, join.get("summaryRowPolicy"))
    left_columns = normalize_string_list(join.get("leftColumns"), "join.leftColumns", maximum=8, required=True)
    right_columns = normalize_string_list(join.get("rightColumns"), "join.rightColumns", maximum=8, required=True)
    if len(left_columns) != len(right_columns):
        raise SpreadsheetError("spreadsheet_join_invalid", "join 左右键数量必须一致。", blocked=True)
    missing_left = sorted(set(left_columns) - available_fields)
    if missing_left:
        raise SpreadsheetError("spreadsheet_column_not_found", "左表联接字段不存在。", {"columns": missing_left}, blocked=True)
    assert_columns_declared(right_metadata, set(right_columns))
    assert_authoritative_columns(right_metadata, set(right_columns))
    assert_columns_declared(right_metadata, set(parsers.get(right_table_id, {})))
    right_rows = apply_parsers(right_rows, parsers.get(right_table_id, {}))
    join_type = join.get("type", "left")
    cardinality = join.get("cardinality")
    if join_type not in {"left", "inner"} or cardinality not in {"one_to_one", "many_to_one", "one_to_many"}:
        raise SpreadsheetError("spreadsheet_join_invalid", "join.type 或 cardinality 无效。", blocked=True)
    assert_join_keys_complete(left_rows, left_columns, "左表")
    assert_join_keys_complete(right_rows, right_columns, "右表")
    left_keys = [join_key(row, left_columns) for row in left_rows]
    right_keys = [join_key(row, right_columns) for row in right_rows]
    if cardinality in {"one_to_one", "one_to_many"} and has_duplicates(left_keys):
        raise SpreadsheetError("spreadsheet_join_cardinality_failed", "左表联接键不满足声明的唯一性。", {"cardinality": cardinality})
    if cardinality in {"one_to_one", "many_to_one"} and has_duplicates(right_keys):
        raise SpreadsheetError("spreadsheet_join_cardinality_failed", "右表联接键不满足声明的唯一性。", {"cardinality": cardinality})

    prefix = join.get("prefix") or f"join{index + 1}"
    if not re.fullmatch(r"[A-Za-z0-9_-]{1,40}", prefix):
        raise SpreadsheetError("spreadsheet_join_invalid", "join.prefix 格式无效。", blocked=True)
    right_payload_columns = [column for column in right_metadata["headers"] if column not in right_columns]
    formula_aliases = {
        f"{prefix}.{column}"
        for column in right_metadata.get("formulaColumns") or []
        if column in right_payload_columns
    }
    requested_formulas = sorted(formula_aliases & requested_fields)
    if requested_formulas:
        raise SpreadsheetError(
            "spreadsheet_formula_not_authoritative",
            "联接表公式列的缓存值不能作为权威计算输入；请改用基础字段重新计算。",
            {"tableId": right_table_id, "columns": requested_formulas},
            blocked=True,
        )
    aliased_columns = [f"{prefix}.{column}" for column in right_payload_columns]
    collisions = sorted(available_fields & set(aliased_columns))
    if collisions:
        raise SpreadsheetError(
            "spreadsheet_join_column_conflict",
            "联接前缀会覆盖已有字段。",
            {"columns": collisions, "prefix": prefix},
            blocked=True,
        )
    index_map: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for key, row in zip(right_keys, right_rows):
        index_map[key].append(row)
    projected_rows = 0
    for key in left_keys:
        match_count = len(index_map.get(key, []))
        projected_rows += match_count or (1 if join_type == "left" else 0)
        if projected_rows > MAX_JOIN_OUTPUT_ROWS:
            raise SpreadsheetError(
                "spreadsheet_join_result_too_large",
                "联接结果行数超过安全上限；请先筛选、聚合或修正联接基数。",
                {
                    "maxRows": MAX_JOIN_OUTPUT_ROWS,
                    "projectedRowsAtLeast": projected_rows,
                    "tableId": right_table_id,
                },
                blocked=True,
            )
    output: list[dict[str, Any]] = []
    for left_row, key in zip(left_rows, left_keys):
        matches = index_map.get(key, [])
        if not matches and join_type == "left":
            merged = dict(left_row)
            merged.update({column: None for column in aliased_columns})
            output.append(merged)
        for right_row in matches:
            merged = dict(left_row)
            for column in right_payload_columns:
                merged[f"{prefix}.{column}"] = right_row.get(column)
            output.append(merged)
    return output, right_table_id, set(aliased_columns)


def apply_filters(rows: list[dict[str, Any]], filters: Any) -> list[dict[str, Any]]:
    if not isinstance(filters, list) or len(filters) > 100:
        raise SpreadsheetError("spreadsheet_filter_invalid", "filters 必须是最多 100 项数组。", blocked=True)
    output = rows
    for condition in filters:
        if not isinstance(condition, dict):
            raise SpreadsheetError("spreadsheet_filter_invalid", "filter 必须是对象。", blocked=True)
        column = require_name(condition.get("column"), "filter.column")
        ensure_columns(output, [column])
        operator = condition.get("operator")
        if operator not in {"eq", "neq", "gt", "gte", "lt", "lte", "in", "not_in", "is_null", "not_null"}:
            raise SpreadsheetError("spreadsheet_filter_invalid", f"不支持的 filter operator: {operator}", blocked=True)
        output = [row for row in output if filter_matches(row.get(column), operator, condition)]
    return output


def filter_matches(value: Any, operator: str, condition: dict[str, Any]) -> bool:
    if operator == "is_null":
        return value in (None, "")
    if operator == "not_null":
        return value not in (None, "")
    if operator in {"in", "not_in"}:
        expected = condition.get("values")
        if not isinstance(expected, list):
            raise SpreadsheetError("spreadsheet_filter_invalid", f"{operator} 需要 values 数组。", blocked=True)
        contains = any(compare_values(value, item) == 0 for item in expected)
        return contains if operator == "in" else not contains
    comparison = compare_values(value, condition.get("value"))
    return {
        "eq": comparison == 0,
        "neq": comparison != 0,
        "gt": comparison > 0,
        "gte": comparison >= 0,
        "lt": comparison < 0,
        "lte": comparison <= 0,
    }[operator]


def aggregate_rows(rows: list[dict[str, Any]], group_by: list[str], measures: Any, null_policy: str, not_computable: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not isinstance(measures, list) or not measures or len(measures) > 100:
        raise SpreadsheetError("spreadsheet_measure_invalid", "measures 必须是 1-100 项数组。", blocked=True)
    if null_policy not in {"fail", "exclude"}:
        raise SpreadsheetError("spreadsheet_measure_invalid", "nullPolicy 只支持 fail 或 exclude。", blocked=True)
    ensure_columns(rows, group_by)
    normalized_measures: list[tuple[str, str, str | None]] = []
    measure_ids: set[str] = set()
    for measure in measures:
        if not isinstance(measure, dict):
            raise SpreadsheetError("spreadsheet_measure_invalid", "measure 必须是对象。", blocked=True)
        measure_id = require_name(measure.get("id"), "measure.id")
        if measure_id in group_by or measure_id in measure_ids:
            raise SpreadsheetError("spreadsheet_measure_invalid", f"measure.id 重复或与 groupBy 字段冲突: {measure_id}", blocked=True)
        measure_ids.add(measure_id)
        operation = measure.get("operation")
        if operation not in {"sum", "count", "countDistinct", "min", "max", "mean"}:
            raise SpreadsheetError("spreadsheet_measure_invalid", f"不支持的聚合: {operation}", blocked=True)
        column = measure.get("column")
        if operation != "count" or column is not None:
            column = require_name(column, "measure.column")
            ensure_columns(rows, [column])
        normalized_measures.append((measure_id, operation, column))

    groups: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[tuple(row.get(column) for column in group_by)].append(row)
    if not group_by and not groups:
        groups[tuple()] = []
    output: list[dict[str, Any]] = []
    for key, group_rows in groups.items():
        result = {column: value for column, value in zip(group_by, key)}
        for measure_id, operation, column in normalized_measures:
            value = aggregate_measure(group_rows, operation, column, null_policy)
            result[measure_id] = value
        output.append(result)
    return output


def aggregate_measure(rows: list[dict[str, Any]], operation: str, column: str | None, null_policy: str) -> Any:
    if operation == "count" and column is None:
        return len(rows)
    values = [row.get(column) for row in rows]
    missing = [value for value in values if value in (None, "")]
    if missing and null_policy == "fail":
        raise SpreadsheetError(
            "spreadsheet_null_value",
            f"字段 {column} 含缺失值；默认不允许静默排除。",
            {"column": column, "nullCount": len(missing)},
        )
    values = [value for value in values if value not in (None, "")]
    if operation == "count":
        return len(values)
    if operation == "countDistinct":
        return len({canonical_scalar(value) for value in values})
    if not values:
        return None
    if operation in {"sum", "mean"}:
        decimals = [parse_decimal(value, None) for value in values]
        total = sum(decimals, Decimal("0"))
        return total if operation == "sum" else total / Decimal(len(decimals))
    return aggregate_extreme(values, operation, column)


def aggregate_extreme(values: list[Any], operation: str, column: str | None) -> Any:
    numeric_values: list[Decimal] = []
    numeric = True
    for value in values:
        try:
            numeric_values.append(parse_decimal(value, None))
        except SpreadsheetError as error:
            if error.code == "spreadsheet_numeric_format_ambiguous":
                raise
            numeric = False
            break
    if numeric:
        return min(numeric_values) if operation == "min" else max(numeric_values)
    if all(isinstance(value, str) for value in values):
        return min(values) if operation == "min" else max(values)
    if all(isinstance(value, (date, datetime)) for value in values):
        return (min if operation == "min" else max)(values, key=lambda value: value.isoformat())
    raise SpreadsheetError(
        "spreadsheet_value_type_mixed",
        f"字段 {column} 含不可安全比较的混合类型，无法执行 {operation}。",
        {"column": column, "types": sorted({value_type(value) for value in values})},
    )


def apply_derived_metrics(rows: list[dict[str, Any]], derived: Any, not_computable: list[dict[str, Any]], base_fields: set[str] | None = None) -> int:
    if not isinstance(derived, list) or len(derived) > 100:
        raise SpreadsheetError("spreadsheet_derived_invalid", "derivedMetrics 必须是最多 100 项数组。", blocked=True)
    existing_fields = set(base_fields or ()) | (set(rows[0]) if rows else set())
    metric_ids: set[str] = set()
    not_computable_count = 0
    for item in derived:
        if not isinstance(item, dict):
            raise SpreadsheetError("spreadsheet_derived_invalid", "derivedMetric 必须是对象。", blocked=True)
        metric_id = require_name(item.get("id"), "derivedMetric.id")
        if metric_id in metric_ids or metric_id in existing_fields:
            raise SpreadsheetError("spreadsheet_derived_invalid", f"derivedMetric.id 重复或覆盖已有字段: {metric_id}", blocked=True)
        metric_ids.add(metric_id)
        existing_fields.add(metric_id)
        operation = item.get("operation")
        if operation not in {"add", "subtract", "multiply", "divide"}:
            raise SpreadsheetError("spreadsheet_derived_invalid", f"不支持的派生运算: {operation}", blocked=True)
        assert_metric_operand(item.get("left"), existing_fields)
        assert_metric_operand(item.get("right"), existing_fields)
        zero_policy = item.get("zeroPolicy", "not_computable")
        if zero_policy not in {"not_computable", "fail"}:
            raise SpreadsheetError("spreadsheet_derived_invalid", "zeroPolicy 只支持 not_computable 或 fail。", blocked=True)
        scale = item.get("scale")
        if scale is not None and (not isinstance(scale, int) or scale < 0 or scale > 12):
            raise SpreadsheetError("spreadsheet_derived_invalid", "scale 必须是 0-12 的整数。", blocked=True)
        rounding_name = item.get("rounding", "half_up")
        if rounding_name not in {"half_up", "half_even"}:
            raise SpreadsheetError("spreadsheet_derived_invalid", "rounding 只支持 half_up 或 half_even。", blocked=True)
        rounding = ROUND_HALF_UP if rounding_name == "half_up" else ROUND_HALF_EVEN
        for row_index, row in enumerate(rows):
            left = resolve_metric_operand(row, item.get("left"))
            right = resolve_metric_operand(row, item.get("right"))
            if left is None or right is None or (operation == "divide" and right == 0):
                if operation == "divide" and right == 0 and zero_policy == "fail":
                    raise SpreadsheetError("spreadsheet_zero_denominator", f"派生指标 {metric_id} 的分母为零。")
                row[metric_id] = None
                not_computable_count += 1
                if len(not_computable) < MAX_NOT_COMPUTABLE_SAMPLES:
                    not_computable.append({"metric": metric_id, "rowIndex": row_index, "reason": "zero_denominator" if right == 0 else "missing_operand"})
                continue
            value = {
                "add": left + right,
                "subtract": left - right,
                "multiply": left * right,
                "divide": left / right,
            }[operation]
            if scale is not None:
                value = value.quantize(Decimal(1).scaleb(-scale), rounding=rounding)
            row[metric_id] = value
    return not_computable_count


def assert_metric_operand(operand: Any, available_fields: set[str]) -> None:
    if isinstance(operand, (int, float, Decimal)) and not isinstance(operand, bool):
        return
    if isinstance(operand, str):
        if operand in available_fields:
            return
        try:
            parse_decimal(operand, None)
            return
        except SpreadsheetError as error:
            raise SpreadsheetError(
                "spreadsheet_derived_invalid",
                f"派生指标引用了不存在的字段或无效常量: {operand}",
                blocked=True,
            ) from error
    raise SpreadsheetError("spreadsheet_derived_invalid", f"派生指标操作数无效: {operand!r}", blocked=True)


def validate_spreadsheet(workspace: Path, arguments: dict[str, Any]) -> dict[str, Any]:
    analysis_id = require_id(arguments.get("analysisId"), "analysisId", "analysis")
    manifest, analysis_directory, source_paths = load_analysis(workspace, analysis_id)
    ensure_sources_unchanged(manifest, source_paths)
    result_ids = normalize_string_list(arguments.get("resultIds"), "resultIds", maximum=100)
    results = {result_id: load_result(analysis_directory, analysis_id, result_id) for result_id in result_ids}
    checks = arguments.get("checks")
    if not isinstance(checks, list) or not checks or len(checks) > MAX_CHECKS:
        raise SpreadsheetError("spreadsheet_checks_invalid", f"checks 必须是 1-{MAX_CHECKS} 项数组。", blocked=True)
    context = WorkbookContext(source_paths, manifest)
    check_ids: set[str] = set()
    outcomes: list[dict[str, Any]] = []
    try:
        for check in checks:
            if not isinstance(check, dict):
                raise SpreadsheetError("spreadsheet_check_invalid", "check 必须是对象。", blocked=True)
            check_id = require_name(check.get("id"), "check.id")
            if check_id in check_ids:
                raise SpreadsheetError("spreadsheet_check_invalid", f"check.id 重复: {check_id}", blocked=True)
            check_ids.add(check_id)
            check_type = check.get("type")
            if "required" in check and not isinstance(check["required"], bool):
                raise SpreadsheetError("spreadsheet_check_invalid", "check.required 必须是布尔值。", blocked=True)
            required = check.get("required", True)
            if check_type == "column_quality":
                outcome = validate_column_quality(context, results, check)
            elif check_type == "numeric_compare":
                outcome = validate_numeric_compare(results, check)
            elif check_type == "set_relation":
                outcome = validate_set_relation(results, check)
            elif check_type == "source_coverage":
                outcome = validate_source_coverage(manifest, results, check)
            else:
                raise SpreadsheetError("spreadsheet_check_invalid", f"不支持的 check.type: {check_type}", blocked=True)
            outcome.update({"id": check_id, "type": check_type, "required": required})
            outcomes.append(outcome)
    finally:
        context.close()

    ensure_sources_unchanged(manifest, source_paths)
    required_failures = [item for item in outcomes if item["required"] and item["status"] == "failed"]
    warnings = [item for item in outcomes if not item["required"] and item["status"] == "failed"]
    validation_status = "failed" if required_failures else "passed_with_warnings" if warnings else "passed"
    validation_id = f"validation-{uuid.uuid4().hex}"
    report = {
        "schemaVersion": "agent-spreadsheet.validation.v1",
        "analysisId": analysis_id,
        "validationId": validation_id,
        "createdAt": iso_now(),
        "sourceSetHash": manifest["sourceSetHash"],
        "resultIds": result_ids,
        "validationStatus": validation_status,
        "counts": {
            "total": len(outcomes),
            "passed": sum(1 for item in outcomes if item["status"] == "passed"),
            "failed": sum(1 for item in outcomes if item["status"] == "failed"),
            "requiredFailed": len(required_failures),
            "warnings": len(warnings),
        },
        "checks": outcomes,
    }
    if len(manifest["sources"]) == 1:
        report["sourceHash"] = manifest["sources"][0]["sha256"]
    validation_directory = ensure_safe_directory(analysis_directory, "validations")
    report_path = validation_directory / f"{validation_id}.json"
    atomic_write_json(report_path, report)
    public_checks, checks_truncated = compact_validation_outcomes(outcomes)
    return {
        "analysisId": analysis_id,
        "validationId": validation_id,
        "validationStatus": validation_status,
        "counts": report["counts"],
        "checks": public_checks,
        "checksTruncated": checks_truncated,
        "reportPath": relative_posix(workspace, report_path),
    }


def compact_validation_outcomes(outcomes: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], bool]:
    """优先返回失败证据；完整检查始终保存在 validation report 中。"""
    prioritized = [item for item in outcomes if item.get("status") == "failed"]
    prioritized.extend(item for item in outcomes if item.get("status") != "failed")
    selected = prioritized[:MAX_PUBLIC_VALIDATION_CHECKS]
    compacted: list[dict[str, Any]] = []
    for item in selected:
        value, evidence_truncated = compact_public_evidence(item)
        if evidence_truncated and isinstance(value, dict):
            value["evidenceTruncated"] = True
        compacted.append(value)
    return compacted, len(outcomes) > len(selected) or any(item.get("evidenceTruncated") for item in compacted)


def compact_public_evidence(value: Any, depth: int = 0) -> tuple[Any, bool]:
    if depth > 8:
        return "[证据层级已截断]", True
    if isinstance(value, str):
        if len(value) <= MAX_PUBLIC_EVIDENCE_STRING_CHARS:
            return value, False
        return f"{value[:MAX_PUBLIC_EVIDENCE_STRING_CHARS]}…", True
    if isinstance(value, list):
        items = value[:MAX_PUBLIC_EVIDENCE_ITEMS]
        output = []
        truncated = len(value) > len(items)
        for item in items:
            compacted, item_truncated = compact_public_evidence(item, depth + 1)
            output.append(compacted)
            truncated = truncated or item_truncated
        return output, truncated
    if isinstance(value, dict):
        entries = list(value.items())[:MAX_PUBLIC_EVIDENCE_FIELDS]
        output: dict[str, Any] = {}
        truncated = len(value) > len(entries)
        for key, item in entries:
            compacted, item_truncated = compact_public_evidence(item, depth + 1)
            output[key] = compacted
            truncated = truncated or item_truncated
        return output, truncated
    return value, False


def validate_column_quality(context: WorkbookContext, results: dict[str, dict[str, Any]], check: dict[str, Any]) -> dict[str, Any]:
    target = check.get("target")
    if not isinstance(target, dict):
        raise SpreadsheetError("spreadsheet_check_invalid", "column_quality.target 必须是对象。", blocked=True)
    if target.get("tableId"):
        metadata, rows = context.table(target["tableId"])
        available = set(metadata.get("headers") or [])
    elif target.get("resultId"):
        result = require_result(results, target["resultId"])
        rows = result["data"]["rows"]
        available = {item.get("name") for item in result["data"].get("columns") or [] if item.get("name")}
    else:
        raise SpreadsheetError("spreadsheet_check_invalid", "column_quality.target 需要 tableId 或 resultId。", blocked=True)
    violations: list[dict[str, Any]] = []
    minimum_rows = check.get("minimumRows", 0)
    if not isinstance(minimum_rows, int) or minimum_rows < 0:
        raise SpreadsheetError("spreadsheet_check_invalid", "minimumRows 必须是非负整数。", blocked=True)
    if len(rows) < minimum_rows:
        violations.append({"rule": "minimumRows", "expected": minimum_rows, "actual": len(rows)})
    specs = check.get("columns") or []
    if not isinstance(specs, list) or len(specs) > 100:
        raise SpreadsheetError("spreadsheet_check_invalid", "column_quality.columns 必须是最多 100 项数组。", blocked=True)
    for spec in specs:
        if not isinstance(spec, dict):
            raise SpreadsheetError("spreadsheet_check_invalid", "column quality 列规则必须是对象。", blocked=True)
        name = require_name(spec.get("name"), "column rule.name")
        if name not in available:
            if spec.get("required", True) is not False:
                violations.append({"column": name, "rule": "required", "message": "字段不存在"})
            continue
        values = [row.get(name) for row in rows]
        if spec.get("notNull") is True:
            null_indexes = [index for index, value in enumerate(values) if value in (None, "")]
            if null_indexes:
                violations.append({"column": name, "rule": "notNull", "count": len(null_indexes), "sampleRows": null_indexes[:10]})
        minimum_coverage = spec.get("minCoverage")
        if minimum_coverage is not None:
            if not isinstance(minimum_coverage, (int, float)) or isinstance(minimum_coverage, bool) or not 0 <= minimum_coverage <= 1:
                raise SpreadsheetError("spreadsheet_check_invalid", "minCoverage 必须是 0-1 的数值。", blocked=True)
            populated = sum(1 for value in values if value not in (None, ""))
            coverage = populated / len(values) if values else 0
            if coverage < minimum_coverage:
                violations.append({
                    "column": name,
                    "rule": "minCoverage",
                    "expected": str(minimum_coverage),
                    "actual": format(coverage, ".6f"),
                    "populated": populated,
                    "rowCount": len(values),
                })
        if spec.get("unique") is True:
            counts = Counter(canonical_scalar(value) for value in values if value not in (None, ""))
            duplicates = [value for value, count in counts.items() if count > 1]
            if duplicates:
                violations.append({"column": name, "rule": "unique", "count": len(duplicates), "samples": duplicates[:10]})
        expected_type = spec.get("type")
        if expected_type:
            invalid = [index for index, value in enumerate(values) if value not in (None, "") and not matches_type(value, expected_type)]
            if invalid:
                violations.append({"column": name, "rule": "type", "expected": expected_type, "count": len(invalid), "sampleRows": invalid[:10]})
        forbidden_pattern = spec.get("forbiddenPattern")
        if forbidden_pattern is not None:
            if not isinstance(forbidden_pattern, str) or not forbidden_pattern or len(forbidden_pattern) > 200:
                raise SpreadsheetError("spreadsheet_check_invalid", "forbiddenPattern 必须是 1-200 字符。", blocked=True)
            assert_safe_pattern(forbidden_pattern)
            try:
                pattern = re.compile(forbidden_pattern)
            except re.error as error:
                raise SpreadsheetError("spreadsheet_check_invalid", f"forbiddenPattern 无效: {error}", blocked=True) from error
            matches = [str(value) for value in values if value not in (None, "") and pattern.search(str(value))]
            if matches:
                violations.append({"column": name, "rule": "forbiddenPattern", "count": len(matches), "samples": matches[:10]})
    return {"status": "failed" if violations else "passed", "violations": violations, "rowCount": len(rows)}


def validate_numeric_compare(results: dict[str, dict[str, Any]], check: dict[str, Any]) -> dict[str, Any]:
    left = evaluate_expression(results, check.get("left"))
    right = evaluate_expression(results, check.get("right"))
    operator = check.get("operator", "eq")
    if operator not in {"eq", "lte", "gte"}:
        raise SpreadsheetError("spreadsheet_check_invalid", "numeric_compare.operator 只支持 eq/lte/gte。", blocked=True)
    absolute_tolerance = parse_decimal(check.get("absoluteTolerance", "0"), None)
    relative_tolerance = parse_decimal(check.get("relativeTolerance", "0"), None)
    if absolute_tolerance < 0 or relative_tolerance < 0:
        raise SpreadsheetError("spreadsheet_check_invalid", "数值容差不能为负数。", blocked=True)
    difference = left - right
    absolute_difference = abs(difference)
    relative_to_left = absolute_difference / abs(left) if left != 0 else None
    relative_to_right = absolute_difference / abs(right) if right != 0 else None
    largest_magnitude = max(abs(left), abs(right))
    relative_difference = absolute_difference / largest_magnitude if largest_magnitude != 0 else Decimal("0")
    tolerance = max(absolute_tolerance, abs(right) * relative_tolerance)
    if operator == "eq":
        passed = absolute_difference <= tolerance
    elif operator == "lte":
        passed = left <= right + tolerance
    else:
        passed = left + tolerance >= right
    return {
        "status": "passed" if passed else "failed",
        "operator": operator,
        "left": serialize_value(left),
        "right": serialize_value(right),
        "difference": serialize_value(difference),
        "absoluteDifference": serialize_value(absolute_difference),
        "relativeDifference": serialize_value(relative_difference),
        "relativeToLeft": serialize_value(relative_to_left),
        "relativeToRight": serialize_value(relative_to_right),
        "absoluteTolerance": serialize_value(absolute_tolerance),
        "relativeTolerance": serialize_value(relative_tolerance),
    }


def assert_safe_pattern(pattern: str) -> None:
    # 质量门只需要字段形态校验；禁止高级扩展、反向引用和嵌套量词，避免模型提交
    # 可能造成灾难性回溯的正则拖死 worker。
    nested_quantifier = re.search(
        r"\([^)]*(?:\*|\+|\{\d+(?:,\d*)?\})[^)]*\)(?:\*|\+|\{\d+(?:,\d*)?\})",
        pattern,
    )
    repeated_quantifier = re.search(
        r"(?:\*|\+|\?|\{\d+(?:,\d*)?\})(?:\*|\+|\?|\{\d+(?:,\d*)?\})",
        pattern,
    )
    if "(?" in pattern or re.search(r"\\[1-9]", pattern) or nested_quantifier or repeated_quantifier:
        raise SpreadsheetError(
            "spreadsheet_check_invalid",
            "forbiddenPattern 只能使用基础正则，不能包含高级扩展、反向引用或嵌套量词。",
            blocked=True,
        )


def validate_set_relation(results: dict[str, dict[str, Any]], check: dict[str, Any]) -> dict[str, Any]:
    left = evaluate_set_operand(results, check.get("left"))
    right = evaluate_set_operand(results, check.get("right"))
    relation = check.get("relation")
    if relation not in {"subset", "equal"}:
        raise SpreadsheetError("spreadsheet_check_invalid", "set_relation.relation 只支持 subset/equal。", blocked=True)
    missing = sorted(left - right)
    extra = sorted(right - left)
    passed = not missing if relation == "subset" else not missing and not extra
    return {
        "status": "passed" if passed else "failed",
        "relation": relation,
        "leftCount": len(left),
        "rightCount": len(right),
        "missingFromRight": missing[:100],
        "extraInRight": extra[:100] if relation == "equal" else [],
        "truncated": len(missing) > 100 or (relation == "equal" and len(extra) > 100),
    }


def validate_source_coverage(manifest: dict[str, Any], results: dict[str, dict[str, Any]], check: dict[str, Any]) -> dict[str, Any]:
    allowed = {
        "id",
        "type",
        "required",
        "resultIds",
        "requireAllSourcesAccountedFor",
        "allowFailedSources",
        "allowUnresolvedDuplicates",
        "allowUnresolvedOverlaps",
    }
    if set(check) - allowed:
        raise SpreadsheetError("spreadsheet_check_invalid", "source_coverage 包含不支持字段。", blocked=True)
    selected_ids = normalize_string_list(check.get("resultIds"), "source_coverage.resultIds", maximum=100)
    if not selected_ids:
        selected_ids = list(results)
    selected = [require_result(results, result_id) for result_id in selected_ids]
    used_source_ids: set[str] = set()
    used_table_ids: set[str] = set()
    decisions: dict[str, dict[str, Any]] = {}
    for result in selected:
        lineage = result["manifest"].get("lineage") or {}
        used_source_ids.update(
            source.get("sourceId")
            for source in lineage.get("sources") or []
            if source.get("sourceId")
        )
        used_table_ids.update(result["manifest"].get("tableIds") or [])
        for decision in result["manifest"].get("sourceDecisions") or lineage.get("sourceDecisions") or []:
            if decision.get("sourceId"):
                previous = decisions.get(decision["sourceId"])
                if previous and previous != decision:
                    raise SpreadsheetError(
                        "spreadsheet_source_decisions_conflict",
                        "被校验结果包含互相冲突的来源决定。",
                        {"sourceId": decision["sourceId"], "decisions": [previous, decision]},
                        blocked=True,
                    )
                decisions[decision["sourceId"]] = decision
    known_source_ids = {source["sourceId"] for source in manifest["sources"]}
    excluded_source_ids = {
        source_id for source_id, decision in decisions.items()
        if decision.get("action") == "exclude"
    }
    included_but_unused = sorted(
        source_id for source_id, decision in decisions.items()
        if decision.get("action") == "include" and source_id not in used_source_ids
    )
    excluded_but_used = sorted(excluded_source_ids & used_source_ids)
    unaccounted = sorted(known_source_ids - used_source_ids - excluded_source_ids)
    unresolved_duplicates = [
        group for group in manifest.get("duplicateGroups") or []
        if (
            len(set(group.get("sourceIds") or []) & used_source_ids) > 1
            if group.get("type") == "identical_file"
            else len(set(group.get("tableIds") or []) & used_table_ids) > 1
        )
    ]
    unresolved_overlaps = [
        candidate for candidate in manifest.get("overlapCandidates") or []
        if set(candidate.get("tableIds") or []) <= used_table_ids
    ]
    failed_sources = [
        source for source in manifest.get("failedSources") or []
        if source.get("sourceId") not in excluded_source_ids
    ]
    require_all = require_optional_boolean(check, "requireAllSourcesAccountedFor", True)
    allow_failed = require_optional_boolean(check, "allowFailedSources", False)
    allow_duplicates = require_optional_boolean(check, "allowUnresolvedDuplicates", False)
    allow_overlaps = require_optional_boolean(check, "allowUnresolvedOverlaps", False)
    violations = []
    if require_all and unaccounted:
        violations.append({"code": "unaccounted_sources", "sourceIds": unaccounted})
    if included_but_unused:
        violations.append({"code": "included_sources_not_used", "sourceIds": included_but_unused})
    if excluded_but_used:
        violations.append({"code": "excluded_sources_used", "sourceIds": excluded_but_used})
    if failed_sources and not allow_failed:
        violations.append({"code": "failed_sources", "sources": failed_sources})
    if unresolved_duplicates and not allow_duplicates:
        violations.append({"code": "unresolved_duplicates", "groups": unresolved_duplicates})
    if unresolved_overlaps and not allow_overlaps:
        violations.append({"code": "unresolved_overlaps", "candidates": unresolved_overlaps})
    return {
        "status": "failed" if violations else "passed",
        "resultIds": selected_ids,
        "knownSourceCount": len(known_source_ids),
        "usedSourceIds": sorted(used_source_ids),
        "excludedSourceIds": sorted(excluded_source_ids),
        "unaccountedSourceIds": unaccounted,
        "violations": violations,
    }


def require_optional_boolean(value: dict[str, Any], name: str, default: bool) -> bool:
    if name not in value:
        return default
    if not isinstance(value[name], bool):
        raise SpreadsheetError("spreadsheet_check_invalid", f"{name} 必须是布尔值。", blocked=True)
    return value[name]


def evaluate_expression(results: dict[str, dict[str, Any]], expression: Any, depth: int = 0) -> Decimal:
    if depth > 12 or not isinstance(expression, dict):
        raise SpreadsheetError("spreadsheet_expression_invalid", "数值表达式结构无效。", blocked=True)
    if "value" in expression:
        return parse_decimal(expression["value"], None)
    if "resultId" in expression:
        result = require_result(results, expression["resultId"])
        row_index = expression.get("rowIndex", 0)
        field = require_name(expression.get("field"), "expression.field")
        rows = result["data"]["rows"]
        if not isinstance(row_index, int) or row_index < 0 or row_index >= len(rows) or field not in rows[row_index]:
            raise SpreadsheetError("spreadsheet_expression_invalid", "结果单元格引用无效。", {"resultId": expression["resultId"], "rowIndex": row_index, "field": field}, blocked=True)
        return parse_decimal(rows[row_index][field], None)
    operation = expression.get("operation")
    operands = expression.get("operands")
    if operation not in {"sum", "add", "subtract", "multiply", "divide"} or not isinstance(operands, list) or not operands:
        raise SpreadsheetError("spreadsheet_expression_invalid", "表达式 operation/operands 无效。", blocked=True)
    values = [evaluate_expression(results, operand, depth + 1) for operand in operands]
    if operation in {"subtract", "divide"} and len(values) != 2:
        raise SpreadsheetError("spreadsheet_expression_invalid", f"{operation} 必须恰好有两个操作数。", blocked=True)
    if operation in {"sum", "add"}:
        return sum(values, Decimal("0"))
    if operation == "subtract":
        return values[0] - values[1]
    if operation == "multiply":
        result = Decimal("1")
        for value in values:
            result *= value
        return result
    if values[1] == 0:
        raise SpreadsheetError("spreadsheet_zero_denominator", "校验表达式分母为零。")
    return values[0] / values[1]


def evaluate_set_operand(results: dict[str, dict[str, Any]], operand: Any) -> set[str]:
    if not isinstance(operand, dict):
        raise SpreadsheetError("spreadsheet_set_operand_invalid", "集合操作数必须是对象。", blocked=True)
    if "values" in operand:
        if not isinstance(operand["values"], list):
            raise SpreadsheetError("spreadsheet_set_operand_invalid", "values 必须是数组。", blocked=True)
        return {str(value) for value in operand["values"] if value not in (None, "")}
    result = require_result(results, operand.get("resultId"))
    field = require_name(operand.get("field"), "set operand.field")
    if result["data"]["rows"] and field not in result["data"]["rows"][0]:
        raise SpreadsheetError("spreadsheet_set_operand_invalid", f"结果字段不存在: {field}", blocked=True)
    return {str(row[field]) for row in result["data"]["rows"] if row.get(field) not in (None, "")}


def load_analysis(workspace: Path, analysis_id: str) -> tuple[dict[str, Any], Path, dict[str, Path]]:
    expected_directory = workspace / "temp" / "spreadsheets" / analysis_id
    try:
        analysis_directory = expected_directory.resolve(strict=True)
    except FileNotFoundError as error:
        raise SpreadsheetError("spreadsheet_analysis_not_found", "找不到表格分析记录。", blocked=True) from error
    ensure_child_path(workspace, analysis_directory)
    if not analysis_directory.is_dir():
        raise SpreadsheetError("spreadsheet_analysis_invalid", "表格分析路径不是目录。", blocked=True)
    manifest_path = analysis_directory / "manifest.json"
    manifest = read_json(manifest_path, "spreadsheet_analysis_not_found", "找不到表格分析记录。")
    if manifest.get("schemaVersion") not in {"agent-spreadsheet.analysis.v1", "agent-spreadsheet.analysis.v2"} or manifest.get("analysisId") != analysis_id:
        raise SpreadsheetError("spreadsheet_analysis_invalid", "表格分析 manifest 无效。", blocked=True)
    manifest = normalize_analysis_manifest(manifest)
    source_paths = {
        source["sourceId"]: resolve_source_path(workspace, source.get("path"))
        for source in manifest["sources"]
    }
    return manifest, analysis_directory, source_paths


def normalize_analysis_manifest(manifest: dict[str, Any]) -> dict[str, Any]:
    if manifest.get("schemaVersion") == "agent-spreadsheet.analysis.v2":
        if not isinstance(manifest.get("sources"), list) or not manifest["sources"]:
            raise SpreadsheetError("spreadsheet_analysis_invalid", "多来源分析缺少 sources。", blocked=True)
        return manifest
    source = dict(manifest.get("source") or {})
    if not source.get("path") or not source.get("sha256"):
        raise SpreadsheetError("spreadsheet_analysis_invalid", "旧版分析缺少 source。", blocked=True)
    source_id = create_source_id(source["path"], source["sha256"])
    source.update({
        "sourceId": source_id,
        "status": "ready",
        "sheetCount": len(manifest.get("sheets") or []),
        "tableCount": len(manifest.get("tables") or []),
    })
    normalized = dict(manifest)
    normalized["sourceSetHash"] = source["sha256"]
    normalized["sources"] = [source]
    normalized["duplicateGroups"] = []
    normalized["overlapCandidates"] = []
    normalized["failedSources"] = []
    normalized["sheets"] = [
        {**sheet, "sourceId": source_id, "sourcePath": source["path"]}
        for sheet in manifest.get("sheets") or []
    ]
    normalized["tables"] = [
        {**table, "sourceId": source_id, "sourcePath": source["path"]}
        for table in manifest.get("tables") or []
    ]
    return normalized


def table_lineage(manifest: dict[str, Any], table_id: str) -> dict[str, Any]:
    table = next((item for item in manifest.get("tables") or [] if item.get("tableId") == table_id), None)
    if table is None:
        raise SpreadsheetError("spreadsheet_table_not_found", f"找不到 tableId: {table_id}", blocked=True)
    return {
        "tableId": table_id,
        "sourceId": table.get("sourceId"),
        "sourcePath": table.get("sourcePath"),
        "sheet": table.get("sheet"),
        "range": table.get("range"),
        "kind": table.get("kind"),
        "declaredName": table.get("declaredName"),
    }


def result_source_lineage(manifest: dict[str, Any], table_ids: list[str]) -> list[dict[str, Any]]:
    source_ids = {
        table.get("sourceId")
        for table in manifest.get("tables") or []
        if table.get("tableId") in table_ids
    }
    return [
        {
            "sourceId": source["sourceId"],
            "path": source["path"],
            "sha256": source["sha256"],
        }
        for source in manifest["sources"]
        if source["sourceId"] in source_ids
    ]


def load_result(analysis_directory: Path, analysis_id: str, result_id: str) -> dict[str, Any]:
    require_id(result_id, "resultId", "result")
    try:
        manifest_path = (analysis_directory / "results" / f"{result_id}.manifest.json").resolve(strict=True)
    except FileNotFoundError as error:
        raise SpreadsheetError("spreadsheet_result_not_found", f"找不到 resultId: {result_id}", blocked=True) from error
    ensure_child_path(analysis_directory, manifest_path)
    manifest = read_json(manifest_path, "spreadsheet_result_not_found", f"找不到 resultId: {result_id}")
    if manifest.get("schemaVersion") != "agent-spreadsheet.result.v1" or manifest.get("analysisId") != analysis_id or manifest.get("resultId") != result_id:
        raise SpreadsheetError("spreadsheet_result_invalid", f"结果 manifest 无效: {result_id}", blocked=True)
    data_file = manifest.get("dataFile")
    if not isinstance(data_file, str) or not data_file:
        raise SpreadsheetError("spreadsheet_result_invalid", f"结果 manifest 缺少 dataFile: {result_id}", blocked=True)
    try:
        data_path = (analysis_directory / data_file).resolve(strict=True)
    except FileNotFoundError as error:
        raise SpreadsheetError("spreadsheet_result_not_found", f"找不到 resultId 数据文件: {result_id}", blocked=True) from error
    ensure_child_path(analysis_directory.resolve(strict=True), data_path)
    if hash_file(data_path) != manifest.get("dataHash"):
        raise SpreadsheetError("spreadsheet_result_hash_mismatch", f"结果内容哈希不一致: {result_id}")
    data = read_json(data_path, "spreadsheet_result_invalid", f"无法读取结果数据: {result_id}")
    return {"manifest": manifest, "data": data}


def require_result(results: dict[str, dict[str, Any]], result_id: Any) -> dict[str, Any]:
    if not isinstance(result_id, str) or result_id not in results:
        raise SpreadsheetError("spreadsheet_result_not_allowed", f"结果未列入 resultIds: {result_id}", blocked=True)
    result = results[result_id]
    if result.get("manifest", {}).get("truncated"):
        raise SpreadsheetError(
            "spreadsheet_result_truncated",
            f"受限结果不能作为完整质量门证据: {result_id}",
            {"resultId": result_id, "rowCount": result["manifest"].get("rowCount"), "returnedRows": result["manifest"].get("returnedRows")},
            blocked=True,
        )
    return result


def ensure_sources_unchanged(manifest: dict[str, Any], source_paths: dict[str, Path]) -> None:
    changed = []
    for source in manifest["sources"]:
        source_id = source["sourceId"]
        source_path = source_paths[source_id]
        current_hash = hash_file(source_path)
        if current_hash != source["sha256"]:
            changed.append({
                "sourceId": source_id,
                "path": source["path"],
                "expectedHash": source["sha256"],
                "actualHash": current_hash,
            })
    if changed:
        raise SpreadsheetError(
            "spreadsheet_source_changed",
            "来源集合在 inspect 后发生变化，请重新执行 spreadsheet_inspect。",
            {"sources": changed},
            blocked=True,
        )


def assert_authoritative_columns(metadata: dict[str, Any], columns: set[str]) -> None:
    formulas = set(metadata.get("formulaColumns") or [])
    invalid = sorted(formulas & columns)
    if invalid:
        raise SpreadsheetError(
            "spreadsheet_formula_not_authoritative",
            "公式列的缓存值不能作为权威计算输入；请改用基础字段重新计算。",
            {"tableId": metadata["tableId"], "columns": invalid},
            blocked=True,
        )


def assert_columns_declared(metadata: dict[str, Any], columns: set[str]) -> None:
    missing = sorted(columns - set(metadata.get("headers") or []))
    if missing:
        raise SpreadsheetError(
            "spreadsheet_column_not_found",
            "表格字段不存在。",
            {"tableId": metadata["tableId"], "columns": missing},
            blocked=True,
        )


def normalized_join_prefixes(joins: Any) -> set[str]:
    prefixes: set[str] = set()
    for index, join in enumerate(joins if isinstance(joins, list) else []):
        if not isinstance(join, dict):
            continue
        prefix = join.get("prefix") or f"join{index + 1}"
        if prefix in prefixes:
            raise SpreadsheetError("spreadsheet_join_invalid", f"join.prefix 重复: {prefix}", blocked=True)
        prefixes.add(prefix)
    return prefixes


def query_input_fields(query: dict[str, Any]) -> set[str]:
    fields = set(query.get("columns") or []) | set(query.get("groupBy") or [])
    for measure in query.get("measures") or []:
        if isinstance(measure, dict) and measure.get("column"):
            fields.add(measure["column"])
    for condition in query.get("filters") or []:
        if isinstance(condition, dict) and condition.get("column"):
            fields.add(condition["column"])
    return fields


def join_left_fields(joins: Any) -> set[str]:
    if not isinstance(joins, list):
        return set()
    return {
        field
        for join in joins
        if isinstance(join, dict) and isinstance(join.get("leftColumns"), list)
        for field in join["leftColumns"]
        if isinstance(field, str)
    }


def sort_fields(sort_spec: Any) -> set[str]:
    if not isinstance(sort_spec, list):
        return set()
    return {
        item["field"]
        for item in sort_spec
        if isinstance(item, dict) and isinstance(item.get("field"), str)
    }


def referenced_columns(query: dict[str, Any], table_id: str, join_prefixes: set[str], base_headers: set[str]) -> set[str]:
    columns = {
        field for field in query_input_fields(query)
        if field in base_headers
        or not any(field.startswith(f"{prefix}.") for prefix in join_prefixes)
    }
    for parser in query.get("columnParsers") or []:
        if isinstance(parser, dict) and (parser.get("tableId") or table_id) == table_id and parser.get("column"):
            columns.add(parser["column"])
    for join in query.get("joins") or []:
        if isinstance(join, dict):
            columns.update(
                field
                for field in join.get("leftColumns") or []
                if isinstance(field, str) and (
                    field in base_headers
                    or not any(field.startswith(f"{prefix}.") for prefix in join_prefixes)
                )
            )
    return columns


def assert_join_keys_complete(rows: list[dict[str, Any]], columns: list[str], side: str) -> None:
    missing_count = sum(
        1 for row in rows
        if any(row.get(column) in (None, "") for column in columns)
    )
    if missing_count:
        raise SpreadsheetError(
            "spreadsheet_join_key_null",
            f"{side}联接键存在空值，不能安全联接。",
            {"columns": columns, "rowCount": missing_count},
        )


def parse_typed_value(value: Any, parser: dict[str, Any]) -> Any:
    parser_type = parser["type"]
    if value in (None, ""):
        return None
    if parser_type == "string":
        return str(value)
    if parser_type == "decimal":
        return parse_decimal(value, parser)
    if parser_type == "integer":
        decimal = parse_decimal(value, parser)
        if decimal != decimal.to_integral_value():
            raise SpreadsheetError("spreadsheet_value_parse_failed", f"值不是整数: {value}")
        return int(decimal)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    text = str(value).strip()
    try:
        return datetime.fromisoformat(text).isoformat()
    except ValueError as error:
        raise SpreadsheetError("spreadsheet_value_parse_failed", f"无法解析日期: {value}") from error


def parse_decimal(value: Any, parser: dict[str, Any] | None) -> Decimal:
    if value is None or value == "":
        raise SpreadsheetError("spreadsheet_value_not_computable", "缺失值不能参与数值计算。")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        raise SpreadsheetError("spreadsheet_value_parse_failed", "布尔值不能参与数值计算。")
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if parser:
        for symbol in parser.get("currencySymbols") or []:
            text = text.replace(str(symbol), "")
        thousands = parser.get("thousandsSeparator")
        decimal_separator = parser.get("decimalSeparator", ".")
        if thousands:
            text = text.replace(thousands, "")
        if decimal_separator != ".":
            text = text.replace(decimal_separator, ".")
    elif "," in text:
        raise SpreadsheetError(
            "spreadsheet_numeric_format_ambiguous",
            f"数值 {value!r} 含逗号，必须通过 columnParsers 明确小数和千位分隔符。",
        )
    text = text.replace(" ", "")
    if not re.fullmatch(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)", text):
        raise SpreadsheetError("spreadsheet_value_parse_failed", f"无法解析数值: {value!r}")
    try:
        return Decimal(text)
    except InvalidOperation as error:
        raise SpreadsheetError("spreadsheet_value_parse_failed", f"无法解析数值: {value!r}") from error


def resolve_metric_operand(row: dict[str, Any], operand: Any) -> Decimal | None:
    if isinstance(operand, (int, float, Decimal)):
        return parse_decimal(operand, None)
    if isinstance(operand, str):
        if operand in row:
            return None if row[operand] is None else parse_decimal(row[operand], None)
        return parse_decimal(operand, None)
    raise SpreadsheetError("spreadsheet_derived_invalid", f"派生指标操作数无效: {operand!r}", blocked=True)


def compare_values(left: Any, right: Any) -> int:
    try:
        left_value = parse_decimal(left, None)
        right_value = parse_decimal(right, None)
    except SpreadsheetError:
        left_value = "" if left is None else str(left)
        right_value = "" if right is None else str(right)
    return (left_value > right_value) - (left_value < right_value)


def sort_rows(rows: list[dict[str, Any]], sort_spec: Any, available_fields: set[str]) -> list[dict[str, Any]]:
    if not isinstance(sort_spec, list) or len(sort_spec) > 20:
        raise SpreadsheetError("spreadsheet_sort_invalid", "sort 必须是最多 20 项数组。", blocked=True)
    output = list(rows)
    for spec in reversed(sort_spec):
        if not isinstance(spec, dict):
            raise SpreadsheetError("spreadsheet_sort_invalid", "sort 条目必须是对象。", blocked=True)
        field = require_name(spec.get("field"), "sort.field")
        if field not in available_fields:
            raise SpreadsheetError("spreadsheet_column_not_found", f"排序字段不存在: {field}", blocked=True)
        direction = spec.get("direction", "asc")
        if direction not in {"asc", "desc"}:
            raise SpreadsheetError("spreadsheet_sort_invalid", "sort.direction 只支持 asc/desc。", blocked=True)
        output.sort(key=lambda row: sort_key(row.get(field)), reverse=direction == "desc")
    return output


def sort_key(value: Any) -> tuple[int, int, Any]:
    if value is None:
        return (1, 0, "")
    try:
        return (0, 0, Decimal(str(value)))
    except InvalidOperation:
        return (0, 1, str(value))


def infer_output_columns(rows: list[dict[str, Any]], expected_names: list[str] | None = None) -> list[dict[str, str]]:
    names: list[str] = list(expected_names or [])
    for row in rows:
        for name in row:
            if name not in names:
                names.append(name)
    return [{"name": name, "type": infer_serial_type([row.get(name) for row in rows])} for name in names]


def infer_serial_type(values: list[Any]) -> str:
    non_null = [value for value in values if value is not None]
    if not non_null:
        return "unknown"
    if all(isinstance(value, int) and not isinstance(value, bool) for value in non_null):
        return "integer"
    if all(isinstance(value, (int, float, Decimal)) and not isinstance(value, bool) for value in non_null):
        return "decimal"
    if all(isinstance(value, bool) for value in non_null):
        return "boolean"
    return "string"


def matches_type(value: Any, expected: str) -> bool:
    if expected == "string":
        return isinstance(value, str)
    if expected == "integer":
        try:
            number = parse_decimal(value, None)
            return number == number.to_integral_value()
        except SpreadsheetError:
            return False
    if expected in {"decimal", "number"}:
        try:
            parse_decimal(value, None)
            return True
        except SpreadsheetError:
            return False
    if expected == "date":
        if isinstance(value, (date, datetime)):
            return True
        try:
            datetime.fromisoformat(str(value))
            return True
        except ValueError:
            return False
    if expected == "boolean":
        return isinstance(value, bool)
    raise SpreadsheetError("spreadsheet_check_invalid", f"不支持的字段类型: {expected}", blocked=True)


def connected_ranges(cells: set[tuple[int, int]]) -> list[tuple[int, int, int, int]]:
    remaining = set(cells)
    ranges: list[tuple[int, int, int, int]] = []
    while remaining:
        start = remaining.pop()
        queue = deque([start])
        component = [start]
        while queue:
            row, col = queue.popleft()
            for neighbor in ((row - 1, col), (row + 1, col), (row, col - 1), (row, col + 1)):
                if neighbor in remaining:
                    remaining.remove(neighbor)
                    queue.append(neighbor)
                    component.append(neighbor)
        rows = [item[0] for item in component]
        columns = [item[1] for item in component]
        ranges.append((min(rows), min(columns), max(rows), max(columns)))
    return sorted(ranges)


def normalize_headers(values: list[Any]) -> tuple[list[str], list[str]]:
    headers: list[str] = []
    warnings: list[str] = []
    for index, value in enumerate(values):
        if value in (None, ""):
            headers.append(f"column_{index + 1}")
            warnings.append(f"第 {index + 1} 列表头为空。")
        else:
            headers.append(str(value).strip())
    duplicates = sorted(name for name, count in Counter(headers).items() if count > 1)
    if duplicates:
        warnings.append(f"表头重复: {', '.join(duplicates)}")
    return headers, warnings


def ensure_safe_archive(source_path: Path) -> None:
    try:
        with zipfile.ZipFile(source_path) as archive:
            entries = archive.infolist()
            if not entries:
                raise SpreadsheetError("spreadsheet_file_invalid", "工作簿 ZIP 不包含任何文件。", blocked=True)
            total = sum(item.file_size for item in entries)
            compressed = sum(max(item.compress_size, 1) for item in entries)
            if total > MAX_ARCHIVE_BYTES or total / compressed > MAX_ARCHIVE_RATIO:
                raise SpreadsheetError(
                    "spreadsheet_archive_unsafe",
                    "工作簿解压体积或压缩比超过安全上限。",
                    {"uncompressedBytes": total, "compressionRatio": total / compressed},
                    blocked=True,
                )
    except zipfile.BadZipFile as error:
        raise SpreadsheetError("spreadsheet_file_invalid", "工作簿不是有效的 XLSX/XLSM 文件。", blocked=True) from error


def resolve_workspace(value: Any) -> Path:
    if not isinstance(value, str) or not value.strip():
        raise SpreadsheetError("spreadsheet_workspace_required", "缺少 workspace。", blocked=True)
    workspace = Path(value).resolve()
    if not workspace.is_dir():
        raise SpreadsheetError("spreadsheet_workspace_invalid", "workspace 不存在或不是目录。", blocked=True)
    return workspace


def resolve_source_path(workspace: Path, value: Any) -> Path:
    if not isinstance(value, str) or not value.strip():
        raise SpreadsheetError("spreadsheet_path_required", "path 必须是 workspace 内文件路径。", blocked=True)
    candidate = Path(value)
    if not candidate.is_absolute():
        candidate = workspace / candidate
    try:
        resolved = candidate.resolve(strict=True)
    except FileNotFoundError as error:
        raise SpreadsheetError("spreadsheet_source_not_found", f"找不到表格文件: {value}", blocked=True) from error
    ensure_child_path(workspace, resolved)
    if not resolved.is_file():
        raise SpreadsheetError("spreadsheet_source_invalid", "表格来源不是文件。", blocked=True)
    if resolved.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise SpreadsheetError("spreadsheet_format_unsupported", "只支持 xlsx、xlsm、csv 和 tsv。", {"extension": resolved.suffix.lower()}, blocked=True)
    if resolved.stat().st_size > MAX_SOURCE_BYTES:
        raise SpreadsheetError("spreadsheet_too_large", "表格文件超过 100MB 安全上限。", {"bytes": resolved.stat().st_size}, blocked=True)
    return resolved


def ensure_child_path(parent: Path, child: Path) -> None:
    try:
        child.relative_to(parent)
    except ValueError as error:
        raise SpreadsheetError("spreadsheet_path_escape", "路径必须位于当前 workspace 内。", blocked=True) from error


def ensure_safe_directory(parent: Path, name: str) -> Path:
    expected = parent / name
    expected.mkdir(exist_ok=True)
    resolved_parent = parent.resolve(strict=True)
    resolved = expected.resolve(strict=True)
    ensure_child_path(resolved_parent, resolved)
    if not resolved.is_dir():
        raise SpreadsheetError("spreadsheet_path_invalid", f"表格存储路径不是目录: {name}", blocked=True)
    return resolved


def read_delimited_rows(path_value: Path, delimiter: str) -> list[list[Any]]:
    try:
        with path_value.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = [list(row) for row in csv.reader(handle, delimiter=delimiter)]
    except UnicodeDecodeError as error:
        raise SpreadsheetError("spreadsheet_encoding_unsupported", "CSV/TSV 必须使用 UTF-8 或 UTF-8 BOM。", blocked=True) from error
    if sum(len(row) for row in rows) > MAX_NON_EMPTY_CELLS:
        raise SpreadsheetError("spreadsheet_too_large", "分隔文本单元格数量超过安全上限。", blocked=True)
    return rows


def write_result_csv(path_value: Path, columns: list[dict[str, Any]], rows: list[dict[str, Any]]) -> None:
    temporary = path_value.with_suffix(path_value.suffix + f".{uuid.uuid4().hex}.tmp")
    names = [column["name"] for column in columns]
    with temporary.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=names, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    os.replace(temporary, path_value)


def atomic_write_json(path_value: Path, value: Any) -> None:
    path_value.parent.mkdir(parents=True, exist_ok=True)
    temporary = path_value.with_suffix(path_value.suffix + f".{uuid.uuid4().hex}.tmp")
    temporary.write_text(json.dumps(value, ensure_ascii=False, indent=2, default=json_default) + "\n", encoding="utf-8")
    os.replace(temporary, path_value)


def read_json(path_value: Path, code: str, message: str) -> dict[str, Any]:
    try:
        return json.loads(path_value.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError) as error:
        raise SpreadsheetError(code, message, blocked=True) from error


def public_table_summary(table: dict[str, Any]) -> dict[str, Any]:
    columns = []
    for column in table["columns"][:50]:
        profile = {
            "name": column["name"],
            "inferredType": column["inferredType"],
            "formulaStatus": column["formulaStatus"],
        }
        for field in ["nullCount", "duplicateValueCount", "ambiguousNumericCount"]:
            if column[field]:
                profile[field] = column[field]
        columns.append(profile)
    summary = {
        "tableId": table["tableId"],
        "sourceId": table.get("sourceId"),
        "sourcePath": table.get("sourcePath"),
        "sheet": table["sheet"],
        "range": table["range"],
        "kind": table["kind"],
        "declaredName": table["declaredName"],
        "rowCount": table["rowCount"],
        "columnCount": table["columnCount"],
        "columns": columns,
        "sampleRows": [compact_row(row, 8) for row in table["sampleRows"][:1]],
    }
    for name, value in {
        "headerWarnings": table["headerWarnings"],
        "formulaColumns": table["formulaColumns"],
        "summaryRows": table.get("summaryRows") or [],
        "mergedRanges": table["mergedRanges"][:20],
    }.items():
        if value:
            summary[name] = value
    if len(table["columns"]) > len(columns):
        summary["columnsTruncated"] = True
    if table["needsSelection"]:
        summary["needsSelection"] = True
    if table.get("structureAmbiguous"):
        summary["structureAmbiguous"] = True
    return summary


def compact_result_preview(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    selected = rows[:MAX_PUBLIC_PREVIEW_ROWS]
    if not selected:
        return []
    maximum_columns = min(
        MAX_PUBLIC_PREVIEW_COLUMNS,
        max(1, MAX_PUBLIC_PREVIEW_CELLS // len(selected)),
    )
    return [compact_row(row, maximum_columns) for row in selected]


def compact_row(row: dict[str, Any], maximum_columns: int) -> dict[str, Any]:
    output: dict[str, Any] = {}
    for index, (key, value) in enumerate(row.items()):
        if index >= maximum_columns:
            output["__columnsTruncated"] = True
            break
        serialized = serialize_value(value)
        if isinstance(serialized, str) and len(serialized) > 200:
            serialized = f"{serialized[:200]}…"
        output[key] = serialized
    return output


def serialize_row(headers: list[str], row: list[Any]) -> dict[str, Any]:
    return {header: serialize_value(row[index] if index < len(row) else None) for index, header in enumerate(headers)}


def serialize_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def json_default(value: Any) -> Any:
    serialized = serialize_value(value)
    if serialized is value:
        return str(value)
    return serialized


def value_type(value: Any) -> str:
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, (float, Decimal)):
        return "decimal"
    if isinstance(value, (datetime, date)):
        return "date"
    return "string"


def infer_type(counts: Counter[str]) -> str:
    if not counts:
        return "unknown"
    return counts.most_common(1)[0][0] if len(counts) == 1 else "mixed"


def canonical_scalar(value: Any) -> str:
    return json.dumps(serialize_value(value), ensure_ascii=False, sort_keys=True, default=json_default)


def looks_ambiguous_numeric(value: str) -> bool:
    stripped = value.strip()
    return bool(re.fullmatch(r"[-+]?[\d., ]+", stripped) and "," in stripped)


def join_key(row: dict[str, Any], columns: list[str]) -> tuple[Any, ...]:
    return tuple(canonical_scalar(row.get(column)) for column in columns)


def has_duplicates(values: list[tuple[Any, ...]]) -> bool:
    return len(values) != len(set(values))


def ensure_columns(rows: list[dict[str, Any]], columns: list[str]) -> None:
    if not columns or not rows:
        return
    available = set(rows[0].keys()) if rows else set()
    missing = [column for column in columns if column not in available]
    if missing:
        raise SpreadsheetError("spreadsheet_column_not_found", "表格字段不存在。", {"columns": missing}, blocked=True)


def require_name(value: Any, field: str) -> str:
    if not isinstance(value, str) or not value.strip() or len(value.strip()) > 120:
        raise SpreadsheetError("spreadsheet_invalid_input", f"{field} 必须是 1-120 字符。", blocked=True)
    return value.strip()


def require_id(value: Any, field: str, prefix: str) -> str:
    normalized = require_name(value, field)
    if not ID_RE.fullmatch(normalized) or not normalized.startswith(f"{prefix}-"):
        raise SpreadsheetError("spreadsheet_invalid_input", f"{field} 格式无效。", blocked=True)
    return normalized


def normalize_string_list(value: Any, field: str, maximum: int, required: bool = False) -> list[str]:
    if value is None and not required:
        return []
    if not isinstance(value, list) or (required and not value) or len(value) > maximum:
        raise SpreadsheetError("spreadsheet_invalid_input", f"{field} 必须是最多 {maximum} 项的字符串数组。", blocked=True)
    return [require_name(item, field) for item in value]


def hash_file(path_value: Path) -> str:
    digest = hashlib.sha256()
    with path_value.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def relative_posix(parent: Path, child: Path) -> str:
    return child.relative_to(parent).as_posix()


def ensure_cell_range_size(min_row: int, min_col: int, max_row: int, max_col: int, label: str) -> None:
    """在读取单元格前拒绝异常大的已用区域，避免稀疏工作簿触发全表扫描。"""
    rows = max_row - min_row + 1
    columns = max_col - min_col + 1
    if columns > MAX_TABLE_COLUMNS:
        raise SpreadsheetError(
            "spreadsheet_too_wide",
            f"{label} 的列数超过安全上限。",
            {"maxColumns": MAX_TABLE_COLUMNS, "columns": columns},
            blocked=True,
        )
    cells = rows * columns
    if cells > MAX_SCANNED_CELLS:
        raise SpreadsheetError(
            "spreadsheet_scan_too_large",
            f"{label} 的声明范围过大，无法安全扫描。",
            {"maxScannedCells": MAX_SCANNED_CELLS, "declaredCells": cells},
            blocked=True,
        )


def ranges_overlap(a_min_row: int, a_min_col: int, a_max_row: int, a_max_col: int, b_min_row: int, b_min_col: int, b_max_row: int, b_max_col: int) -> bool:
    return not (a_max_row < b_min_row or b_max_row < a_min_row or a_max_col < b_min_col or b_max_col < a_min_col)


def iso_now() -> str:
    return datetime.now().astimezone().isoformat(timespec="milliseconds")


if __name__ == "__main__":
    main()
