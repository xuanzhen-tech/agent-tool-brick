"""为 spreadsheet smoke 生成真实 XLSX/XLSM 测试文件。"""

from pathlib import Path
import re
import shutil
import sys
import zipfile

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def add_table(sheet, name, reference):
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)


workspace = Path(sys.argv[1])
uploads = workspace / "uploads"
uploads.mkdir(parents=True, exist_ok=True)

workbook = Workbook()
campaign = workbook.active
campaign.title = "Campaign"
campaign.append(["campaignId", "spend", "sales", "clicks", "orders"])
spend = [5000, 4800, 5200, 4900, 5100, 4700, 5000, 4841.62]
sales = [30000, 28000, 29000, 27000, 26000, 25000, 28000, 25490]
for index, (spend_value, sales_value) in enumerate(zip(spend, sales), start=1):
    campaign.append([f"campaign-{index}", spend_value, sales_value, 10_000 if index < 8 else 10_600, 1_000 if index < 8 else 1_464])
add_table(campaign, "CampaignData", "A1:E9")

partial = workbook.create_sheet("Partial")
partial.append(["rowId", "spend", "sales"])
for row in [
    ["part-1", 8000, 43000],
    ["part-2", 7500, 42000],
    ["part-3", 7000, 41000],
    ["part-4", 7107.13, 41555.52],
]:
    partial.append(row)
add_table(partial, "PartialData", "A1:C5")

keywords = workbook.create_sheet("Keywords")
keywords.append(["keyword", "action"])
keywords.append(["B012345678", "negative"])
keywords.append(["B087654321", "negative"])
add_table(keywords, "KeywordData", "A1:B3")

formulas = workbook.create_sheet("Formula")
formulas.append(["rowId", "base", "formulaTotal"])
formulas.append(["f-1", 10, "=B2*2"])
formulas.append(["f-2", 20, "=B3*2"])
add_table(formulas, "FormulaData", "A1:C3")

mappings = workbook.create_sheet("Mappings")
mappings.append(["campaignId", "owner"])
mappings.append(["campaign-1", "A"])
mappings.append(["campaign-1", "B"])
add_table(mappings, "MappingData", "A1:B3")

owners = workbook.create_sheet("Owners")
owners.append(["campaignId", "owner", "formulaScore"])
owners.append(["campaign-1", "Alice", "=1+1"])
owners.append(["campaign-2", "Bob", "=2+2"])
add_table(owners, "OwnerData", "A1:C3")

null_mappings = workbook.create_sheet("NullMappings")
null_mappings.append(["campaignId", "owner"])
null_mappings.append([None, "Unknown"])
add_table(null_mappings, "NullMappingData", "A1:B2")

totals = workbook.create_sheet("Totals")
totals.append(["rowId", "spend"])
totals.append(["a", 10])
totals.append(["b", 20])
totals.append(["Total", 30])
add_table(totals, "TotalsData", "A1:B4")

empty_rows = workbook.create_sheet("EmptyRows")
empty_rows.append(["rowId", "spend"])
empty_rows.append(["a", 10])
empty_rows.append([None, None])
empty_rows.append(["b", 20])
add_table(empty_rows, "EmptyRowsData", "A1:B4")

zero_metrics = workbook.create_sheet("ZeroMetrics")
zero_metrics.append(["rowId", "spend", "sales"])
zero_metrics.append(["zero", 10, 0])
add_table(zero_metrics, "ZeroMetricsData", "A1:C2")

ambiguous = workbook.create_sheet("Merged")
ambiguous.merge_cells("A1:B1")
ambiguous["A1"] = "广告明细"
ambiguous.append(["id", "value"])
ambiguous.append(["a", 1])
ambiguous.append(["b", 2])

hidden = workbook.create_sheet("Hidden")
hidden.sheet_state = "hidden"
hidden.append(["id", "value"])
hidden.append(["hidden-1", 1])
add_table(hidden, "HiddenData", "A1:B2")

xlsx_path = uploads / "advertising.xlsx"
workbook.save(xlsx_path)
workbook.close()
shutil.copyfile(xlsx_path, uploads / "advertising.xlsm")
with zipfile.ZipFile(uploads / "empty-workbook.xlsx", "w"):
    pass

# 人工写入过期公式缓存，验证 inspect 只能标记 cached_unverified，不能把缓存值
# 当成权威金额。openpyxl 不负责计算公式，因此直接在 OOXML 中构造缓存证据。
with zipfile.ZipFile(xlsx_path, "r") as source_archive, zipfile.ZipFile(uploads / "stale-formula.xlsx", "w") as target_archive:
    replacement_count = 0
    for item in source_archive.infolist():
        payload = source_archive.read(item.filename)
        if item.filename == "xl/worksheets/sheet4.xml":
            payload, count = re.subn(
                rb'(<c r="C[23]"[^>]*><f>.*?</f><v>).*?(</v></c>)',
                rb'\g<1>999\g<2>',
                payload,
            )
            replacement_count += count
        target_archive.writestr(item, payload)
if replacement_count != 2:
    raise RuntimeError(f"无法构造公式缓存 fixture: replacements={replacement_count}")

# 这是压缩比异常的 zip，不是合法工作簿；worker 应在交给 openpyxl 前阻断。
with zipfile.ZipFile(uploads / "unsafe-archive.xlsx", "w", compression=zipfile.ZIP_DEFLATED) as archive:
    archive.writestr("xl/worksheets/sheet1.xml", "0" * (2 * 1024 * 1024))

# 文件本身体积很小，但声明的已用区域极大。worker 必须在逐单元格扫描前拒绝，
# 避免稀疏工作簿造成 CPU 和内存放大。
sparse_workbook = Workbook()
sparse_sheet = sparse_workbook.active
sparse_sheet.title = "Sparse"
sparse_sheet["A1"] = "id"
sparse_sheet.cell(row=100_000, column=100, value="tail")
sparse_workbook.save(uploads / "sparse-dimensions.xlsx")
sparse_workbook.close()
