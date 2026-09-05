"""生成表格 Agent A/B 评测所需的真实 XLSX 数据集。"""

from __future__ import annotations

import json
import shutil
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def add_table(sheet, name: str, reference: str) -> None:
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)


def save_book(path: Path, sheets: list[tuple[str, list[list[object]], str | None]]) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    for name, rows, table_name in sheets:
        sheet = workbook.create_sheet(name)
        for row in rows:
            sheet.append(row)
        if table_name and rows:
            last_column = column_letter(len(rows[0]))
            add_table(sheet, table_name, f"A1:{last_column}{len(rows)}")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def column_letter(value: int) -> str:
    output = ""
    while value:
        value, remainder = divmod(value - 1, 26)
        output = chr(65 + remainder) + output
    return output


def write_case(root: Path, case_id: str) -> Path:
    case_root = root / case_id
    (case_root / "uploads").mkdir(parents=True, exist_ok=True)
    return case_root


def build_clean_single(root: Path) -> dict:
    case_root = write_case(root, "clean-single-table")
    save_book(case_root / "uploads" / "orders.xlsx", [(
        "Orders",
        [
            ["orderId", "revenue", "refund"],
            ["order-1", 100.10, 0],
            ["order-2", 200.20, 20],
            ["order-3", 50.05, 0],
            ["order-4", 149.65, 10],
        ],
        "OrdersData",
    )])
    return case(
        "clean-single-table",
        "简单",
        ["uploads/orders.xlsx"],
        "计算订单数、revenue 总额、refund 总额和净收入 revenue-refund，并在正式回答前核对恒等式。",
        [number("500"), number("30"), number("470"), regex(r"4\s*(?:笔|个|条|行|orders?)", 15)],
        ["spreadsheet_inspect", "spreadsheet_compute", "spreadsheet_validate"],
    )


def build_reconciliation(root: Path) -> dict:
    case_root = write_case(root, "multi-sheet-reconciliation")
    workbook = Workbook()
    detail = workbook.active
    detail.title = "Detail"
    for row in [
        ["campaignId", "spend"],
        ["A", 1000.10],
        ["B", 2000.20],
        ["C", 3000.30],
    ]:
        detail.append(row)
    add_table(detail, "CampaignDetail", "A1:B4")
    summary = workbook.create_sheet("Summary")
    summary.append(["metric", "reportedValue", "formulaValue"])
    summary.append(["spend", 5900.50, "=SUM(Detail!B2:B4)"])
    add_table(summary, "ReportedSummary", "A1:C2")
    workbook.save(case_root / "uploads" / "campaign-report.xlsx")
    return case(
        "multi-sheet-reconciliation",
        "中等",
        ["uploads/campaign-report.xlsx"],
        "从 Detail 基础行重算 spend，与 Summary 的 reportedValue 对账。公式列不能作为权威值；给出两边金额和绝对差额，不一致时明确数据未闭环。",
        [number("6000.6"), number("5900.5"), number("100.1"), regex(r"未闭环|不一致|对账失败", 25)],
        ["spreadsheet_inspect", "spreadsheet_compute", "spreadsheet_validate"],
    )


def build_localized_union(root: Path) -> dict:
    case_root = write_case(root, "multi-file-localized-union")
    july = case_root / "uploads" / "july.xlsx"
    save_book(july, [(
        "Campaign",
        [
            ["date", "campaignId", "spend", "sales"],
            ["2026-07-01", "A", 10.10, 100],
            ["2026-07-02", "B", 20.20, 200],
        ],
        "JulyCampaigns",
    )])
    shutil.copy2(july, case_root / "uploads" / "july-copy.xlsx")
    save_book(case_root / "uploads" / "august.xlsx", [(
        "广告活动",
        [
            ["日期", "广告活动", "花费", "销售额"],
            ["2026-08-01", "C", 30.30, 300],
            ["2026-08-02", "A", 5.05, 50],
        ],
        "AugustCampaigns",
    )])
    save_book(case_root / "uploads" / "owners.xlsx", [(
        "Owners",
        [["campaignId", "owner"], ["A", "team-1"], ["B", "team-2"], ["C", "team-1"]],
        "OwnerMapping",
    )])
    (case_root / "uploads" / "broken.xlsx").write_text("not an xlsx", encoding="utf8")
    return case(
        "multi-file-localized-union",
        "困难",
        [
            "uploads/july.xlsx",
            "uploads/july-copy.xlsx",
            "uploads/august.xlsx",
            "uploads/owners.xlsx",
            "uploads/broken.xlsx",
        ],
        "识别并排除完全重复的 July 副本和损坏文件；统一中英文字段，合并有效月份并关联 owner。分别报告 team-1、team-2 的 spend/sales，并证明所有来源均已使用或有理由排除。",
        [
            regex(r"team-1", 8), number("45.45", 14), number("450", 10),
            regex(r"team-2", 8), number("20.2", 14), number("200", 10),
            regex(r"重复|副本|完全(?:相同|一致)|duplicate", 8), regex(r"损坏|无法读取|broken|corrupt", 8), regex(r"覆盖|来源|source", 10),
        ],
        ["spreadsheet_inspect", "spreadsheet_compute", "spreadsheet_validate"],
    )


def build_join_trap(root: Path) -> dict:
    case_root = write_case(root, "join-cardinality-trap")
    save_book(case_root / "uploads" / "sales.xlsx", [(
        "Sales",
        [["productId", "revenue"], ["P1", 100], ["P2", 200]],
        "SalesData",
    )])
    save_book(case_root / "uploads" / "categories.xlsx", [(
        "Categories",
        [["productId", "category"], ["P1", "A"], ["P1", "A-duplicate"], ["P2", "B"]],
        "CategoryMapping",
    )])
    return case(
        "join-cardinality-trap",
        "困难",
        ["uploads/sales.xlsx", "uploads/categories.xlsx"],
        "尝试按 category 汇总 revenue，但映射表应满足 many-to-one。先检查联接基数；若 productId 不唯一，必须阻断分类汇总并说明金额可能被放大，不能给出伪造分类结果。",
        [regex(r"阻断|不能|无法|block(?:ed)?|cannot", 35), regex(r"重复|不唯一|基数|many.to.one|not unique", 35), regex(r"放大|重复计算|虚增|扩张|inflate|amplif", 20)],
        ["spreadsheet_inspect", "spreadsheet_compute"],
    )


def build_localized_numbers(root: Path) -> dict:
    case_root = write_case(root, "localized-number-format")
    save_book(case_root / "uploads" / "locale.xlsx", [(
        "Finance",
        [
            ["sku", "gross", "cost"],
            ["A", "1.234,56", "234,56"],
            ["B", "2.000,00", "500,00"],
        ],
        "LocalizedFinance",
    )])
    return case(
        "localized-number-format",
        "中等",
        ["uploads/locale.xlsx"],
        "这些数字使用点号作千位、逗号作小数。显式按该格式解析，计算 gross、cost 和 margin=gross-cost 总额，并校验恒等式；不得把 1.234,56 猜成其他数值。",
        [number("3234.56", 25), number("734.56", 25), number("2500", 25), regex(r"逗号|小数|千位|格式", 15)],
        ["spreadsheet_inspect", "spreadsheet_compute", "spreadsheet_validate"],
    )


def build_large(root: Path) -> dict:
    case_root = write_case(root, "large-workbook-aggregation")
    regions = ["North", "South", "East", "West"]
    totals = {region: 0 for region in regions}
    workbook = Workbook(write_only=False)
    sheet = workbook.active
    sheet.title = "Transactions"
    sheet.append(["rowId", "region", "amount"])
    row_count = 40_000
    for index in range(1, row_count + 1):
        region = regions[(index - 1) % len(regions)]
        amount = (index % 100) + 1
        totals[region] += amount
        sheet.append([f"row-{index}", region, amount])
    add_table(sheet, "TransactionsData", f"A1:C{row_count + 1}")
    workbook.save(case_root / "uploads" / "large.xlsx")
    checks = [regex(r"40[,，]?000|40000", 10)]
    for region in regions:
        checks.extend([regex(region, 5), number(str(totals[region]), 12)])
    checks.append(number(str(sum(totals.values())), 22))
    return case(
        "large-workbook-aggregation",
        "压力",
        ["uploads/large.xlsx"],
        "按 region 汇总 40,000 行 amount，并报告四个 region 的合计与总计。正式回答前核对分组合计等于总计，不得把明细全量打印到工具结果。",
        checks,
        ["spreadsheet_inspect", "spreadsheet_compute", "spreadsheet_validate"],
    )


def case(case_id: str, difficulty: str, files: list[str], task: str, checks: list[dict], required_tools: list[str]) -> dict:
    return {
        "id": case_id,
        "difficulty": difficulty,
        "files": files,
        "task": task,
        "checks": checks,
        "specializedRequiredTools": required_tools,
    }


def number(value: str, weight: int = 15) -> dict:
    return {"type": "number", "value": value, "weight": weight}


def regex(pattern: str, weight: int) -> dict:
    return {"type": "regex", "pattern": pattern, "weight": weight}


def main() -> None:
    output = Path(sys.argv[1]).resolve()
    output.mkdir(parents=True, exist_ok=False)
    cases = [
        build_clean_single(output),
        build_reconciliation(output),
        build_localized_union(output),
        build_join_trap(output),
        build_localized_numbers(output),
        build_large(output),
    ]
    manifest = {
        "schemaVersion": "agent-spreadsheet.eval-set.v1",
        "description": "同模型 run_shell 与专属表格工具 A/B 评测集",
        "cases": cases,
    }
    (output / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf8")
    print(output / "manifest.json")


if __name__ == "__main__":
    main()
